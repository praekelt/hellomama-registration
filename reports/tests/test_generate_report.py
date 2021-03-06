import pytz
import responses
import os
import openpyxl

try:
    import mock
except ImportError:
    from unittest import mock

from datetime import datetime
from django.conf import settings
from django.contrib.auth.models import User
from django.core import mail
from django.db.models.signals import post_save
from django.test import TestCase, override_settings
from openpyxl import load_workbook

from rest_hooks.models import model_saved
from seed_services_client import IdentityStoreApiClient

from registrations.models import (
    Registration, Source, registration_post_save, fire_created_metric,
    fire_unique_operator_metric, fire_message_type_metric, fire_source_metric,
    fire_receiver_type_metric, fire_language_metric, fire_state_metric,
    fire_role_metric)
from ..utils import ExportWorkbook, generate_random_filename
from ..tasks.detailed_report import generate_report
from ..models import ReportTaskStatus


class mockobj(object):
    @classmethod
    def choice(cls, li):
        return li[0]


@override_settings(
    IDENTITY_STORE_URL='http://idstore.example.com/',
    IDENTITY_STORE_TOKEN='idstoretoken',
    MESSAGE_SENDER_URL='http://ms.example.com/',
    MESSAGE_SENDER_TOKEN='mstoken',
    STAGE_BASED_MESSAGING_URL='http://sbm.example.com/',
    STAGE_BASED_MESSAGING_TOKEN='sbmtoken')
class GenerateReportTest(TestCase):
    def setUp(self):
        self.adminuser = User.objects.create()
        self.source = Source.objects.create(
            name='test_source', user=self.adminuser, authority='hw_full')

        def has_listeners(class_name):
            return post_save.has_listeners(class_name)
        assert has_listeners(Registration), (
            "Registration model has no post_save listeners. Make sure"
            " helpers cleaned up properly in earlier tests.")
        post_save.disconnect(receiver=registration_post_save,
                             sender=Registration)
        post_save.disconnect(receiver=fire_created_metric,
                             sender=Registration)
        post_save.disconnect(receiver=fire_source_metric,
                             sender=Registration)
        post_save.disconnect(receiver=fire_unique_operator_metric,
                             sender=Registration)
        post_save.disconnect(receiver=fire_message_type_metric,
                             sender=Registration)
        post_save.disconnect(receiver=fire_receiver_type_metric,
                             sender=Registration)
        post_save.disconnect(receiver=fire_language_metric,
                             sender=Registration)
        post_save.disconnect(receiver=fire_state_metric,
                             sender=Registration)
        post_save.disconnect(receiver=fire_role_metric,
                             sender=Registration)
        post_save.disconnect(receiver=model_saved,
                             dispatch_uid='instance-saved-hook')
        assert not has_listeners(Registration), (
            "Registration model still has post_save listeners. Make sure"
            " helpers cleaned up properly in earlier tests.")

    def tearDown(self):
        def has_listeners(class_name):
            return post_save.has_listeners(class_name)
        assert not has_listeners(Registration), (
            "Registration model still has post_save listeners. Make sure"
            " helpers removed them properly in earlier tests.")
        post_save.connect(receiver=registration_post_save,
                          sender=Registration)
        post_save.connect(receiver=fire_created_metric,
                          sender=Registration)
        post_save.connect(receiver=fire_source_metric,
                          sender=Registration)
        post_save.connect(receiver=fire_unique_operator_metric,
                          sender=Registration)
        post_save.connect(receiver=fire_language_metric,
                          sender=Registration)
        post_save.connect(receiver=fire_state_metric,
                          sender=Registration)
        post_save.connect(receiver=fire_role_metric,
                          sender=Registration)

        post_save.connect(receiver=model_saved,
                          dispatch_uid='instance-saved-hook')

        try:
            with mock.patch('random.choice', mockobj.choice):
                filename = generate_random_filename()
                os.remove(filename)
        except OSError:
            pass

    def assertSheetRow(self, file_name, sheet_name, row_number, expected):
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        rows = list(sheet.rows)
        self.assertEqual(
            [cell.value for cell in rows[row_number]],
            expected)

    def add_registrations(self, num=1, msg_receiver='msg_receiver'):
        for i in range(num):
            Registration.objects.create(
                mother_id='mother_id',
                created_at='2016-01-02 00:00:00+00:00', data={
                    'operator_id': 'operator_id',
                    'receiver_id': 'receiver_id',
                    'gravida': 'gravida',
                    'msg_type': 'msg_type',
                    'last_period_date': 'last_period_date',
                    'language': 'language',
                    'msg_receiver': msg_receiver,
                    'voice_days': 'voice_days',
                    'voice_times': 'voice_times',
                    'preg_week': 'preg_week',
                    'reg_type': 'reg_type',
                },
                validated=True,
                source_id=self.source.id
            )

    def add_identity_callback(
            self, identity='operator_id', linked_id='linked-to-identity-id',
            address='+2340000000000'):
        responses.add(
            responses.GET,
            'http://idstore.example.com/identities/{}/'.format(identity),
            json={
                'identity': identity,
                'details': {
                    'personnel_code': 'personnel_code',
                    'facility_name': 'facility_name',
                    'default_addr_type': 'msisdn',
                    'receiver_role': 'role',
                    'state': 'state',
                    'addresses': {
                        'msisdn': {
                            address: {}
                        }
                    },
                    'linked_to': linked_id,
                }
            },
            status=200,
            content_type='application/json')

        if linked_id is not None:
            responses.add(
                responses.GET,
                'http://idstore.example.com/identities/{}/'.format(linked_id),
                json={
                    'identity': linked_id,
                    'details': {
                        'addresses': {
                            'msisdn': {
                                '+2340000000001': {},
                            },
                        },
                    },
                },
                status=200,
                content_type='application/json')

    def add_identity_address_callback(
            self, identity='operator_id', msisdn="+27711445511"):
        responses.add(
            responses.GET,
            'http://idstore.example.com/identities/{}/addresses/msisdn?default=True'.format(identity),  # noqa
            match_querystring=True,
            json={
                "next": None,
                "previous": None,
                "results": [
                    {
                        "address": msisdn
                    }
                ]
            },
            status=200,
            content_type='application/json')

    def add_blank_subscription_callback(self, next_='?foo=bar'):
        if next_:
            next_ = 'http://sbm.example.com/subscriptions/{}'.format(next_)

        responses.add(
            responses.GET,
            ("http://sbm.example.com/subscriptions/?"
             "created_before=2016-02-01T23%3A59%3A59.999999%2B00%3A00"),
            match_querystring=True,
            json={
                'next': next_,
                'results': [],
            },
            status=200,
            content_type='application/json')

    def add_subscriptions_callback(self, path='?foo=bar', num=1):
        subscriptions = [{
            'lang': 'eng_NG',
            'created_at': '2016-11-22T08:12:45.343829Z',
            'messageset': 4,
            'schedule': 5,
            'url': 'url',
            'completed': False,
            'initial_sequence_number': 1,
            'updated_at': '2016-11-22T08:12:52.411545Z',
            'version': 1,
            'next_sequence_number': 1,
            'process_status': 0,
            'active': True,
            'id': '10176584-2a47-42b6-b9f3-a3a98070f35e',
            'identity': '17cf37cf-edd6-4634-88e3-f793575f7e3a',
            'metadata': {
                'scheduler_schedule_id':
                    'a64d153f-1515-42c1-997a-9a3444c916fc'
            }
        }] * num

        responses.add(
            responses.GET,
            'http://sbm.example.com/subscriptions/{}'.format(path),
            match_querystring=True,
            json={
                'next': None,
                'results': subscriptions,
            },
            status=200,
            content_type='application/json')

    def add_blank_outbound_callback(self, next_='?foo=bar'):
        if next_:
            next_ = 'http://ms.example.com/outbound/{}'.format(next_)

        responses.add(
            responses.GET,
            ("http://ms.example.com/outbound/?"
             "before=2016-02-01T23%3A59%3A59.999999%2B00%3A00"
             "&after=2016-01-01T00%3A00%3A00%2B00%3A00"),
            match_querystring=True,
            json={
                'next': next_,
                'results': [],
            },
            status=200,
            content_type='application/json')

    def add_outbound_callback(
            self, num=1, metadata={}, options=(('addr', ''), )):
        outbounds = []

        for i in range(0, num):
            for option in options:
                outbounds.append({
                    'to_addr': option[0],
                    'to_identity': option[1],
                    'content': 'content',
                    'delivered': True if i % 2 == 0 else False,
                    'created_at': '2016-01-01T10:30:21.{}Z'.format(i),
                    'metadata': metadata
                })

        responses.add(
            responses.GET,
            'http://ms.example.com/outbound/?foo=bar',
            match_querystring=True,
            json={
                'next': None,
                'results': outbounds,
            },
            status=200,
            content_type='application/json')

    def add_messageset_callback(self):
        responses.add(
            responses.GET,
            'http://sbm.example.com/messageset/4/',
            json={
                'created_at': '2016-06-22T10:30:21.186435Z',
                'short_name': 'prebirth.mother.audio.10_42.tue_thu.9_11',
                'next_set': 11,
                'notes': '',
                'updated_at': '2016-09-13T13:01:32.591754Z',
                'default_schedule': 5,
                'content_type': 'audio',
                'id': 4
            },
            status=200,
            content_type='application/json')

    def add_blank_optouts_callback(self, next_='?foo=bar'):
        if next_:
            next_ = 'http://idstore.example.com/optouts/search/{}'.format(
                next_)

        responses.add(
            responses.GET,
            ("http://idstore.example.com/optouts/search/?"
             "created_at__lte=2016-02-01T23%3A59%3A59.999999%2B00%3A00&"
             "created_at__gte=2016-01-01T00%3A00%3A00%2B00%3A00"),
            match_querystring=True,
            json={
                'next': next_,
                'results': [],
            },
            status=200,
            content_type='application/json')

    def add_optouts_callback(self, path='?foo=bar', num=1):
        optouts = [{
            "id": "e5210c99-8d8a-40f1-8e7f-8a66c4de9e29",
            "optout_type": "stop",
            "identity": "8311c23d-f3c4-4cab-9e20-5208d77dcd1b",
            "address_type": "msisdn",
            "address": "+1234",
            "request_source": "testsource",
            "requestor_source_id": "1",
            "reason": "Test reason",
            "created_at": "2017-01-27T10:00:06.354178Z"
        }] * num

        responses.add(
            responses.GET,
            'http://idstore.example.com/optouts/search/{}'.format(path),
            match_querystring=True,
            json={
                'next': None,
                'results': optouts,
            },
            status=200,
            content_type='application/json')

    def midnight(self, timestamp):
        return timestamp.replace(hour=0, minute=0, second=0, microsecond=0,
                                 tzinfo=pytz.timezone(settings.TIME_ZONE))

    def trigger_report_generation(self, emails=[]):
        with mock.patch('random.choice', mockobj.choice):
            filename = generate_random_filename()

            task_status = ReportTaskStatus.objects.create(**{
                "start_date": self.midnight(datetime.strptime('2016-01-01',
                                                              '%Y-%m-%d')),
                "end_date": self.midnight(datetime.strptime('2016-02-01',
                                                            '%Y-%m-%d')),
                "email_subject": 'The Email Subject',
                "status": ReportTaskStatus.PENDING
            })

            generate_report.apply_async(kwargs={
                'start_date': self.midnight(datetime.strptime('2016-01-01',
                                                              '%Y-%m-%d')),
                'end_date': self.midnight(datetime.strptime('2016-02-01',
                                                            '%Y-%m-%d')),
                'email_recipients': emails,
                'email_subject': 'The Email Subject',
                'task_status_id': task_status.id})

            return filename

    @responses.activate
    @mock.patch("os.remove")
    def test_generate_report_email(self, mock_remove):
        """
        Generating a report should create an email with the correct address,
        subject, and attachment.
        """
        self.add_blank_subscription_callback(next_=None)
        self.add_blank_outbound_callback(next_=None)
        self.add_blank_optouts_callback(next_=None)
        self.trigger_report_generation(['foo@example.com'])
        [report_email] = mail.outbox
        self.assertEqual(report_email.subject, 'The Email Subject')
        (file_name, data, mimetype) = report_email.attachments[0]
        self.assertEqual('report-2016-01-01-to-2016-02-01.xlsx', file_name)

    @responses.activate
    def test_generate_report_status_done(self):
        """
        Generating a report should mark the ReportTaskStatus objects as Done if
        it is successful.
        """
        self.add_blank_subscription_callback(next_=None)
        self.add_blank_outbound_callback(next_=None)
        self.add_blank_optouts_callback(next_=None)
        self.trigger_report_generation(['foo@example.com'])

        task_status = ReportTaskStatus.objects.last()
        self.assertEqual(task_status.status, ReportTaskStatus.DONE)
        self.assertEqual(task_status.file_size > 7000, True)

    @mock.patch("reports.tasks.send_email.SendEmail.apply_async")
    @responses.activate
    def test_generate_report_status_email(self, mock_send):
        """
        Generating a report should mark the ReportTaskStatus objects as Sending
        before it sends the email.
        """
        self.add_blank_subscription_callback(next_=None)
        self.add_blank_outbound_callback(next_=None)
        self.add_blank_optouts_callback(next_=None)
        self.trigger_report_generation(['foo@example.com'])

        task_status = ReportTaskStatus.objects.last()
        self.assertEqual(task_status.status, ReportTaskStatus.SENDING)
        self.assertEqual(task_status.file_size > 7000, True)

    @responses.activate
    def test_generate_report_status_done_without_email(self):
        """
        Generating a report should mark the ReportTaskStatus objects as Done
        if no email_recipients are specified.
        """

        self.add_blank_subscription_callback(next_=None)
        self.add_blank_outbound_callback(next_=None)
        self.add_blank_optouts_callback(next_=None)
        self.trigger_report_generation()

        task_status = ReportTaskStatus.objects.last()
        self.assertEqual(task_status.status, ReportTaskStatus.DONE)
        self.assertEqual(task_status.file_size > 7000, True)

    @responses.activate
    @mock.patch("reports.tasks.send_email.SendEmail.apply_async")
    def test_generate_report_status_running(self, mock_send):
        """
        Generating a report should mark the ReportTaskStatus objects as Sending
        before it sends the email.
        """

        orig = openpyxl.Workbook.save

        def new_save(wb, file):
            task_status = ReportTaskStatus.objects.last()
            self.assertEqual(task_status.status, ReportTaskStatus.RUNNING)
            orig(wb, file)

        self.add_blank_subscription_callback(next_=None)
        self.add_blank_outbound_callback(next_=None)
        self.add_blank_optouts_callback(next_=None)

        with mock.patch('openpyxl.Workbook.save', new_save):
            self.trigger_report_generation(['foo@example.com'])

    @responses.activate
    @override_settings(
        CELERY_EAGER_PROPAGATES_EXCEPTIONS=False)
    def test_generate_report_status_failed(self):
        """
        Generating a report should mark the ReportTaskStatus objects as Failed
        if there is a problem.
        """
        self.add_blank_subscription_callback(next_=None)
        self.add_blank_outbound_callback(next_=None)
        self.add_blank_optouts_callback(next_=None)

        task_status = ReportTaskStatus.objects.create(**{
            "start_date": "not_really_a_date",
            "end_date": self.midnight(datetime.strptime('2016-02-01',
                                                        '%Y-%m-%d')),
            "email_subject": 'The Email Subject',
            "status": ReportTaskStatus.PENDING
        })

        try:
            generate_report.apply_async(kwargs={
                'start_date': "not_really_a_date",
                'end_date': self.midnight(datetime.strptime('2016-02-01',
                                                            '%Y-%m-%d')),
                'email_recipients': ['foo@example.com'],
                'email_subject': 'The Email Subject',
                'task_status_id': task_status.id})
        except:
            pass

        task_status.refresh_from_db()
        self.assertEqual(task_status.status, ReportTaskStatus.FAILED)
        self.assertEqual(
            task_status.error,
            "time data 'not_really_a_date' does not match format '%Y-%m-%d'")

    @responses.activate
    def test_generate_report_registrations_validated_only(self):
        """
        The registrations report should only return validated registrations.
        """
        self.add_registrations()
        Registration.objects.all().update(validated=False)
        Registration.objects.all().update(created_at='2016-02-01 01:00:00')

        workbook = ExportWorkbook()
        sheet = workbook.add_sheet('testsheet', 0)
        generate_report.handle_registrations(
            sheet, datetime(2016, 2, 1), datetime(2016, 3, 1))

        rows = list(sheet._sheet.rows)
        # Only the header
        self.assertEqual(len(rows), 1)

    @responses.activate
    def test_generate_report_healthworker_registrations_validated_only(self):
        """
        The healthworker registrations report should only return validated
        registrations.
        """
        self.add_registrations()
        Registration.objects.all().update(validated=False)
        Registration.objects.all().update(created_at='2016-02-01 01:00:00')

        workbook = ExportWorkbook()
        sheet = workbook.add_sheet('testsheet', 0)
        generate_report.handle_registrations(
            sheet, datetime(2016, 2, 1), datetime(2016, 3, 1))

        rows = list(sheet._sheet.rows)
        # Only the header
        self.assertEqual(len(rows), 1)

    @responses.activate
    def test_generate_report_registrations_mother_only(self):
        """
        If it is a mother_only registration, then the gatekeeper row should be
        empty.
        """
        self.add_registrations(num=1, msg_receiver='mother_only')
        Registration.objects.all().update(created_at='2016-02-01 01:00:00')

        self.add_identity_callback('operator_id', address='op-addr')
        self.add_identity_callback('receiver_id', address='rec-addr')
        self.add_identity_callback(
            'mother_id', linked_id=None, address='mother-addr')

        generate_report.identity_cache = {}
        generate_report.identity_store_client = IdentityStoreApiClient(
            settings.IDENTITY_STORE_TOKEN,
            settings.IDENTITY_STORE_URL,
        )

        workbook = ExportWorkbook()
        sheet = workbook.add_sheet('testsheet', 0)
        generate_report.handle_registrations(
            sheet, datetime(2016, 2, 1), datetime(2016, 3, 1))

        rows = list(sheet._sheet.rows)
        self.assertEqual(
            [c.value for c in rows[0]],
            [
                'Mother',
                'Gatekeeper',
                'Created',
                'gravida',
                'msg_type',
                'last_period_date',
                'language',
                'msg_receiver',
                'voice_days',
                'Voice_times',
                'preg_week',
                'reg_type',
                'Personnel_code',
                'Facility',
                'Cadre',
                'State',
            ])
        self.assertEqual(
            [c.value for c in rows[1]],
            [
                'mother-addr',
                '',
                '2016-02-01T01:00:00+00:00',
                'gravida',
                'msg_type',
                'last_period_date',
                'language',
                'mother_only',
                'voice_days',
                'voice_times',
                'preg_week',
                'reg_type',
                'personnel_code',
                'facility_name',
                None,
                'state',
            ])

    @responses.activate
    @mock.patch("os.remove")
    def test_generate_report_registrations(self, mock_remove):
        """
        When generating a report, the first tab should be a list of
        registrations with the relevant registration details.
        """
        self.maxDiff = None
        # Add Registrations
        self.add_registrations()
        Registration.objects.all().update(created_at='2016-02-01 01:00:00')

        # HCW Identity
        self.add_identity_callback(address='hcw-addr')

        # Receiver Identity
        self.add_identity_callback('receiver_id', address='receiver-addr')

        # Mother Identity
        self.add_identity_callback('mother_id', address='mother-addr')

        # Subscriptions, first page, just returns empty results to make sure
        # we're actually paging through the results sets using the `next`
        # parameter
        self.add_blank_subscription_callback(next_=None)

        self.add_subscriptions_callback()

        self.add_messageset_callback()

        self.add_identity_callback('17cf37cf-edd6-4634-88e3-f793575f7e3a')

        self.add_blank_outbound_callback(next_=None)

        self.add_outbound_callback()

        # No opt outs, we're not testing optout by subscription
        self.add_blank_optouts_callback(next_=None)

        tmp_file = self.trigger_report_generation(['foo@example.com'])

        mock_remove.assert_called_once_with(tmp_file)

        # Assert headers are set
        self.assertSheetRow(
            tmp_file, 'Registrations by date', 0,
            [
                'Mother',
                'Gatekeeper',
                'Created',
                'gravida',
                'msg_type',
                'last_period_date',
                'language',
                'msg_receiver',
                'voice_days',
                'Voice_times',
                'preg_week',
                'reg_type',
                'Personnel_code',
                'Facility',
                'Cadre',
                'State',
            ])

        # Assert 1 row is written
        self.assertSheetRow(
            tmp_file, 'Registrations by date', 1,
            [
                'mother-addr',
                'receiver-addr',
                '2016-02-01T01:00:00+00:00',
                'gravida',
                'msg_type',
                'last_period_date',
                'language',
                'msg_receiver',
                'voice_days',
                'voice_times',
                'preg_week',
                'reg_type',
                'personnel_code',
                'facility_name',
                None,
                'state',
            ])

    @responses.activate
    @mock.patch("os.remove")
    def test_generate_report_health_worker_registrations(self, mock_remove):
        """
        When generating a report, the second tab should be registrations per
        health worker, and it should have the correct information.
        """
        # Add Registrations and correct date, 2 registrations for 1 operator
        self.add_registrations(num=2)
        Registration.objects.all().update(created_at='2016-01-02 00:00:00')

        # Identity for hcw
        self.add_identity_callback('operator_id')

        # identity for receiver, for first report
        self.add_identity_callback('receiver_id')

        # Mother Identity
        self.add_identity_callback('mother_id')

        # Subscriptions, first page, just returns empty results to make sure
        # we're actually paging through the results sets using the `next`
        # parameter
        self.add_blank_subscription_callback(next_=None)

        self.add_subscriptions_callback(num=2)

        self.add_messageset_callback()

        self.add_identity_callback('17cf37cf-edd6-4634-88e3-f793575f7e3a')

        self.add_blank_outbound_callback(next_=None)

        self.add_outbound_callback()

        # No opt outs, we're not testing optout by subscription
        self.add_blank_optouts_callback(next_=None)

        tmp_file = self.trigger_report_generation(['foo@example.com'])

        mock_remove.assert_called_once_with(tmp_file)

        # Assert headers are set
        self.assertSheetRow(
            tmp_file, 'Health worker registrations', 0,
            [
                'Unique Personnel Code',
                'Facility',
                'State',
                'Cadre',
                'Number of Registrations',
            ])

        # Assert 1 row is written
        self.assertSheetRow(
            tmp_file, 'Health worker registrations', 1,
            [
                'personnel_code',
                'facility_name',
                'state',
                'role',
                2,
            ])

    @responses.activate
    @mock.patch("os.remove")
    def test_generate_report_enrollments(self, mock_remove):
        """
        When generating a report, the third tab should be enrollments,
        and it should have the correct information.
        """
        # Add Registrations, 2 registrations for 1 operator
        self.add_registrations(num=2)

        # Identity for hcw
        self.add_identity_callback('operator_id')

        # identity for receiver, for first report
        self.add_identity_callback('receiver_id')

        # Subscriptions, first page, just returns empty results to make sure
        # we're actually paging through the results sets using the `next`
        # parameter
        self.add_blank_subscription_callback()

        self.add_subscriptions_callback(num=2)

        self.add_messageset_callback()

        self.add_identity_callback('17cf37cf-edd6-4634-88e3-f793575f7e3a')

        self.add_blank_outbound_callback(next_=None)

        self.add_outbound_callback()

        # No opt outs, we're not testing optout by subscription
        self.add_blank_optouts_callback(next_=None)

        tmp_file = self.trigger_report_generation(['foo@example.com'])

        mock_remove.assert_called_once_with(tmp_file)

        # Assert headers are set
        self.assertSheetRow(
            tmp_file, 'Enrollments', 0,
            [
                'Message set',
                'Roleplayer',
                'Total enrolled',
                'Enrolled in period',
                'Enrolled and opted out in period',
                'Enrolled and completed in period',
            ])

        # Assert 1 row is written
        self.assertSheetRow(
            tmp_file, 'Enrollments', 1,
            ['prebirth', 'role', 2, 2, 0, 0])

    @responses.activate
    @mock.patch("os.remove")
    def test_generate_report_sms_per_msisdn(self, mock_remove):
        """
        When generating a report, the fourth tab should be SMS delivery per
        MSISDN, and it should have the correct information.
        """
        # Add Registrations, 2 registrations for 1 operator
        self.add_registrations(num=2)

        # Identity for hcw
        self.add_identity_callback('operator_id')

        # identity for receiver, for first report
        self.add_identity_callback('receiver_id')

        # Subscriptions, first page, just returns empty results to make sure
        # we're actually paging through the results sets using the `next`
        # parameter
        self.add_blank_subscription_callback()

        self.add_subscriptions_callback()

        self.add_messageset_callback()

        self.add_identity_callback('17cf37cf-edd6-4634-88e3-f793575f7e3a')

        self.add_blank_outbound_callback()

        self.add_identity_address_callback('receiver_id', '+2340000000000')

        # Create 4 outbounds with to_addr populated and 4 with to_identity
        self.add_outbound_callback(
            num=4, options=(('+2340000001111', ''), ('', 'receiver_id')))

        # No opt outs, we're not testing optout by subscription
        self.add_blank_optouts_callback(next_=None)

        tmp_file = self.trigger_report_generation(['foo@example.com'])

        mock_remove.assert_called_once_with(tmp_file)

        # Assert headers are set
        self.assertSheetRow(
            tmp_file, 'SMS delivery per MSISDN', 0,
            [
                'MSISDN',
                'SMS 1',
                'SMS 2',
                'SMS 3',
                'SMS 4'
            ])

        # Assert 2 rows are written
        self.assertSheetRow(
            tmp_file, 'SMS delivery per MSISDN', 1,
            ['+2340000000000', 'Yes', 'No', 'Yes', 'No'])
        self.assertSheetRow(
            tmp_file, 'SMS delivery per MSISDN', 2,
            ['+2340000001111', 'Yes', 'No', 'Yes', 'No'])

    @responses.activate
    @mock.patch("os.remove")
    def test_generate_report_obd_delivery_failure(self, mock_remove):
        # Add Registrations, 2 registrations for 1 operator
        self.add_registrations(num=2)

        # Identity for hcw
        self.add_identity_callback('operator_id')

        # identity for receiver, for first report
        self.add_identity_callback('receiver_id')

        # Subscriptions, first page, just returns empty results to make sure
        # we're actually paging through the results sets using the `next`
        # parameter
        self.add_blank_subscription_callback()

        self.add_subscriptions_callback()

        self.add_messageset_callback()

        self.add_identity_callback('17cf37cf-edd6-4634-88e3-f793575f7e3a')

        self.add_blank_outbound_callback()

        self.add_outbound_callback(
            num=40,
            metadata={'voice_speech_url': 'dummy_voice_url'})

        # No opt outs, we're not testing optout by subscription
        self.add_blank_optouts_callback(next_=None)

        tmp_file = self.trigger_report_generation(['foo@example.com'])

        mock_remove.assert_called_once_with(tmp_file)

        # Assert period row
        self.assertSheetRow(
            tmp_file, 'OBD Delivery Failure', 1,
            [
                "In the last period:",
                "2016-01-01 - 2016-02-01",
                None
            ])

        # Check headers
        self.assertSheetRow(
            tmp_file, 'OBD Delivery Failure', 2,
            [
                "OBDs Sent",
                "OBDs failed",
                "Failure rate",
            ]
        )

        # Assert 1 row is written
        self.assertSheetRow(
            tmp_file, 'OBD Delivery Failure', 3,
            [40, 20, '50.00%'])

    @responses.activate
    @mock.patch("os.remove")
    def test_generate_report_optout_by_date(self, mock_remove):
        # Return no registrations or subscriptions for other reports
        self.add_blank_subscription_callback(next_=None)

        # Optouts, first page no results to make sure that we're paging
        self.add_blank_optouts_callback()
        self.add_optouts_callback()

        # Add identity for optout
        self.add_identity_callback('8311c23d-f3c4-4cab-9e20-5208d77dcd1b')

        # Add subscription result for identity
        self.add_subscriptions_callback(
            path=(
                '?active=False&completed=False&'
                'created_before=2017-01-27T10%3A00%3A06.354178Z&'
                'identity=8311c23d-f3c4-4cab-9e20-5208d77dcd1b')
        )

        # Add messageset for subscription
        self.add_messageset_callback()

        self.add_blank_outbound_callback(next_=None)

        tmp_file = self.trigger_report_generation(['foo@example.com'])

        mock_remove.assert_called_once_with(tmp_file)

        # Assert headers are set
        self.assertSheetRow(
            tmp_file, 'Opt Outs by Date', 0,
            [
                "MSISDN",
                "Optout Date",
                "Request Source",
                "Reason"
            ])

        # Assert row 1 is written
        self.assertSheetRow(
            tmp_file, 'Opt Outs by Date', 1,
            [
                "+1234",
                "2017-01-27T10:00:06.354178Z",
                "testsource",
                "Test reason",
            ])

    def test_get_addresses_from_identity(self):
        """
        Getting the addresses from the identity results in the correct
        addresses being returned.
        """
        addresses = generate_report.get_addresses_from_identity({
            'details': {
                'default_addr_type': 'foo',
                'addresses': {
                    'foo': {
                        'addr1': {},
                        'addr2': {},
                    },
                    'bar': {
                        'addr3': {},
                        'addr4': {},
                    },
                },
            },
        })

        self.assertEqual(sorted(addresses), ['addr1', 'addr2'])

    def test_get_addresses_from_identity_defaults_msisdn(self):
        """
        If no default address type is specified, it should default to getting
        msisdn addresses.
        """
        addresses = generate_report.get_addresses_from_identity({
            'details': {
                'addresses': {
                    'msisdn': {
                        'addr1': {},
                        'addr2': {},
                    },
                    'bar': {
                        'addr3': {},
                        'addr4': {},
                    },
                },
            },
        })

        self.assertEqual(sorted(addresses), ['addr1', 'addr2'])

    def test_get_addresses_from_identity_no_msisdns(self):
        """
        If the identity has no addresses of the address type, then an empty
        list should be returned.
        """
        addresses = generate_report.get_addresses_from_identity({
            'details': {
                'addresses': {
                    'bar': {
                        'addr3': {},
                        'addr4': {},
                    },
                },
            },
        })

        self.assertEqual(sorted(addresses), [])

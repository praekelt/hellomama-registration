import collections
import os

from celery.task import Task
from django.conf import settings
from django.core.mail import EmailMessage
from django.utils.dateparse import parse_datetime
from functools import partial
from openpyxl import Workbook
from six import string_types
from datetime import timedelta

from seed_services_client import (IdentityStoreApiClient,
                                  StageBasedMessagingApiClient,
                                  MessageSenderApiClient)

from registrations.models import Registration
from reports.utils import (parse_cursor_params, midnight_validator,
                           generate_random_filename)
from reports.models import ReportTaskStatus


class BaseTask(Task):

    def on_failure(self, exc, task_id, args, kwargs, einfo):
        if kwargs.get('task_status_id'):
            task_status = ReportTaskStatus.objects.get(
                id=kwargs['task_status_id'])

            task_status.status = 'Failed'
            task_status.error = exc
            task_status.save()
            super(BaseTask, self).on_failure(exc, task_id, args, kwargs, einfo)


class SendEmail(BaseTask):

    def run(self, **kwargs):
        subject = kwargs['subject']
        sender = kwargs['sender']
        recipients = kwargs['recipients']
        file_location = kwargs['file_location']
        file_name = kwargs['file_name']
        task_status_id = kwargs['task_status_id']

        email = EmailMessage(subject, '', sender, recipients)
        with open(file_location, 'rb') as fp:
            email.attach(file_name, fp.read(), 'application/vnd.ms-excel')
        email.send()

        task_status = ReportTaskStatus.objects.get(id=task_status_id)
        task_status.status = 'Done'
        task_status.save()

        try:
            os.remove(file_location)
        except OSError:
            pass


class ExportSheet(object):

    def __init__(self, sheet, headers=None):
        self._sheet = sheet
        self.set_header(headers or [])

    def set_header(self, headers, row=1):
        self._headers = headers
        for index, header in enumerate(headers):
            self._sheet.cell(row=row, column=index + 1, value=header)

    def get_header(self):
        return self._headers

    def add_row(self, row):
        row_number = self._sheet.max_row + 1
        for key, value in row.items():
            if isinstance(key, int):
                col_idx = key
            else:
                col_idx = self._headers.index(key) + 1

            cell = self._sheet.cell(
                row=row_number,
                column=col_idx)
            cell.value = value


class ExportWorkbook(object):

    def __init__(self):
        self._workbook = Workbook()

    def add_sheet(self, sheetname, position):
        return ExportSheet(self._workbook.create_sheet(sheetname, position))

    def save(self, file_name):
        return self._workbook.save(file_name)


class GenerateReport(BaseTask):
    workbook_class = ExportWorkbook
    """ Generate an XLS spreadsheet report on registrations, write it to
    disk and email it to specified recipients
    """

    def run(self, start_date, end_date, task_status_id, email_recipients=[],
            email_sender=settings.DEFAULT_FROM_EMAIL,
            email_subject='Seed Control Interface Generated Report', **kwargs):

        task_status = ReportTaskStatus.objects.get(id=task_status_id)
        task_status.status = 'Running'
        task_status.save()

        if isinstance(start_date, string_types):
            start_date = midnight_validator(start_date)
        if isinstance(end_date, string_types):
            end_date = midnight_validator(end_date)

        end_date = end_date + timedelta(days=1, microseconds=-1)

        self.identity_cache = {}
        self.messageset_cache = {}
        self.address_cache = {}

        ids_client = IdentityStoreApiClient(settings.IDENTITY_STORE_TOKEN,
                                            settings.IDENTITY_STORE_URL)
        sbm_client = StageBasedMessagingApiClient(
            settings.STAGE_BASED_MESSAGING_TOKEN,
            settings.STAGE_BASED_MESSAGING_URL)
        ms_client = MessageSenderApiClient(settings.MESSAGE_SENDER_TOKEN,
                                           settings.MESSAGE_SENDER_URL)

        workbook = self.workbook_class()
        sheet = workbook.add_sheet('Registrations by date', 0)
        self.handle_registrations(sheet, ids_client, start_date, end_date)

        sheet = workbook.add_sheet('Health worker registrations', 1)
        self.handle_health_worker_registrations(
            sheet, ids_client, start_date, end_date)

        sheet = workbook.add_sheet('Enrollments', 2)
        self.handle_enrollments(sheet, sbm_client, ids_client, start_date,
                                end_date)

        sheet = workbook.add_sheet('SMS delivery per MSISDN', 3)
        self.handle_sms_delivery_msisdn(sheet, ms_client, ids_client,
                                        start_date, end_date)

        sheet = workbook.add_sheet('OBD Delivery Failure', 4)
        self.handle_obd_delivery_failure(sheet, ms_client, start_date,
                                         end_date)

        sheet = workbook.add_sheet('Opt Outs by Date', 5)
        self.handle_optouts(
            sheet, sbm_client, ids_client, start_date, end_date)

        output_file = generate_random_filename()
        workbook.save(output_file)

        task_status.status = 'Sending'
        task_status.file_size = os.path.getsize(output_file)
        task_status.save()

        if email_recipients:
            file_name = 'report-%s-to-%s.xlsx' % (
                start_date.strftime('%Y-%m-%d'),
                end_date.strftime('%Y-%m-%d'))
            SendEmail.apply_async(kwargs={
                'subject': email_subject,
                'file_name': file_name,
                'file_location': output_file,
                'sender': email_sender,
                'recipients': email_recipients,
                'task_status_id': task_status_id})

    def get_identity(self, ids_client, identity):
        if identity in self.identity_cache:
            return self.identity_cache[identity]

        identity_object = ids_client.get_identity(identity)
        self.identity_cache[identity] = identity_object
        return identity_object

    def get_identity_address(self, ids_client, identity):
        if identity in self.address_cache:
            return self.address_cache[identity]

        address = ids_client.get_identity_address(identity)
        self.address_cache[identity] = address
        return address

    def get_messageset(self, sbm_client, messageset):
        if messageset in self.messageset_cache:
            return self.messageset_cache[messageset]

        messageset_object = sbm_client.get_messageset(messageset)
        self.messageset_cache[messageset] = messageset_object
        return messageset_object

    def get_registrations(self, **kwargs):
        registrations = Registration.objects.filter(**kwargs)
        for result in registrations.iterator():
            yield result

    def get_subscriptions(self, sbm_client, **kwargs):
        subscriptions = sbm_client.get_subscriptions(kwargs)
        cursor = subscriptions['next']
        while cursor:
            for result in subscriptions['results']:
                yield result
            params = parse_cursor_params(cursor)
            subscriptions = sbm_client.get_subscriptions(params)
            cursor = subscriptions['next']
        for result in subscriptions['results']:
            yield result

    def get_outbounds(self, ms_client, **kwargs):
        outbounds = ms_client.get_outbounds(kwargs)
        cursor = outbounds['next']
        while cursor:
            for result in outbounds['results']:
                yield result
            params = parse_cursor_params(cursor)
            outbounds = ms_client.get_outbounds(params)
            cursor = outbounds['next']
        for result in outbounds['results']:
            yield result

    def get_optouts(self, ids_client, **kwargs):
        optouts = ids_client.get_optouts(kwargs)
        cursor = optouts['next']
        while cursor:
            for result in optouts['results']:
                yield result
            params = parse_cursor_params(cursor)
            optouts = ids_client.get_optouts(params)
            cursor = optouts['next']
        for result in optouts['results']:
            yield result

    def handle_registrations(self, sheet, ids_client, start_date, end_date):

        sheet.set_header([
            'MSISDN',
            'Created',
            'gravida',
            'msg_type',
            'last_period_date',
            'language',
            'msg_receiver',
            'voice_days',
            'Voice_times',
            'preg_week',
            'reg_type',
            'Personnel_code',
            'Facility',
            'Cadre',
            'State',
        ])

        registrations = self.get_registrations(
            created_at__gte=start_date.isoformat(),
            created_at__lte=end_date.isoformat())

        for idx, registration in enumerate(registrations):
            data = registration.data
            operator_id = data.get('operator_id')
            receiver_id = data.get('receiver_id')

            if operator_id:
                operator_identity = self.get_identity(
                    ids_client, operator_id)
            else:
                operator_identity = {}

            if receiver_id:
                receiver_identity = self.get_identity(
                    ids_client, receiver_id)
            else:
                receiver_identity = {}

            operator_details = operator_identity.get('details', {})
            receiver_details = receiver_identity.get('details', {})
            default_addr_type = receiver_details.get('default_addr_type')
            if default_addr_type:
                addresses = receiver_details.get('addresses', {})
                msisdns = addresses.get(default_addr_type, {}).keys()
            else:
                msisdns = []

            sheet.add_row({
                'MSISDN': ','.join(msisdns),
                'Created': registration.created_at.isoformat(),
                'gravida': data.get('gravida'),
                'msg_type': data.get('msg_type'),
                'last_period_date': data.get('last_period_date'),
                'language': data.get('language'),
                'msg_receiver': data.get('msg_receiver'),
                'voice_days': data.get('voice_days'),
                'Voice_times': data.get('voice_times'),
                'preg_week': data.get('preg_week'),
                'reg_type': data.get('reg_type'),
                'Personnel_code': operator_details.get('personnel_code'),
                'Facility': operator_details.get('facility_name'),
                'Cadre': operator_details.get('role'),
                'State': operator_details.get('state'),
            })

    def handle_health_worker_registrations(
            self, sheet, ids_client, start_date, end_date):
        sheet.set_header([
            'Unique Personnel Code',
            'Facility',
            'State',
            'Cadre',
            'Number of Registrations'])

        registrations = self.get_registrations(
            created_at__gte=start_date.isoformat(),
            created_at__lte=end_date.isoformat())
        registrations_per_operator = collections.defaultdict(int)

        for registration in registrations:
            operator_id = registration.data.get('operator_id')
            registrations_per_operator[operator_id] += 1

        for operator_id, count in registrations_per_operator.items():
            operator = self.get_identity(ids_client, operator_id) or {}
            operator_details = operator.get('details', {})
            sheet.add_row({
                'Unique Personnel Code': operator_details.get(
                    'personnel_code'),
                'Facility': operator_details.get('facility_name'),
                'State': operator_details.get('state'),
                'Cadre': operator_details.get('receiver_role'),
                'Number of Registrations': count,
            })

    def handle_enrollments(self, sheet, sbm_client, ids_client, start_date,
                           end_date):

        sheet.set_header([
            'Message set',
            'Roleplayer',
            'Total enrolled',
            'Enrolled in period',
            'Enrolled and opted out in period',
            'Enrolled and completed in period',
        ])

        subscriptions = self.get_subscriptions(
            sbm_client,
            created_before=end_date.isoformat())

        data = collections.defaultdict(partial(collections.defaultdict, int))
        for subscription in subscriptions:
            messageset = self.get_messageset(
                sbm_client, subscription['messageset'])
            identity = self.get_identity(ids_client, subscription['identity'])

            messageset_name = messageset['short_name'].split('.')[0]

            receiver_role = 'None'
            if identity:
                receiver_role = identity.get('details', {}).get(
                    'receiver_role', 'None')

            data[messageset_name, receiver_role]['total'] += 1

            if parse_datetime(subscription['created_at']) > start_date:
                data[messageset_name, receiver_role]['total_period'] += 1

                if (not subscription['active'] and
                        not subscription['completed']):
                    data[messageset_name, receiver_role]['optouts'] += 1

                if subscription['completed']:
                    data[messageset_name, receiver_role]['completed'] += 1

        for key in sorted(data.keys()):
            sheet.add_row({
                1: key[0],
                2: key[1],
                3: data[key]['total'],
                4: data[key]['total_period'],
                5: data[key]['optouts'],
                6: data[key]['completed'],
            })

    def handle_sms_delivery_msisdn(
            self, sheet, ms_client, ids_client, start_date, end_date):

        outbounds = self.get_outbounds(
            ms_client,
            after=start_date.isoformat(),
            before=end_date.isoformat()
        )

        data = collections.defaultdict(dict)
        count = collections.defaultdict(int)
        for outbound in outbounds:
            if 'voice_speech_url' not in outbound.get('metadata', {}):

                if (not outbound.get('to_addr', '') and
                        outbound.get('to_identity', '')):
                    outbound['to_addr'] = self.get_identity_address(
                        ids_client, outbound['to_identity'])

                count[outbound['to_addr']] += 1
                data[outbound['to_addr']][outbound['created_at']] = \
                    outbound['delivered']

        if count != {}:
            max_col = max(count.values())

            header = ['MSISDN']
            for col_idx in range(0, max_col):
                header.append('SMS {}'.format(col_idx + 1))

            sheet.set_header(header)

            for msisdn, sms_data in sorted(data.items()):

                row = {1: msisdn}

                for index, (key, state) in enumerate(sorted(sms_data.items())):
                    row[index+2] = 'Yes' if state else 'No'

                sheet.add_row(row)

    def handle_obd_delivery_failure(
            self, sheet, ms_client, start_date, end_date):

        outbounds = self.get_outbounds(
            ms_client,
            after=start_date.isoformat(),
            before=end_date.isoformat()
        )

        data = collections.defaultdict(int)
        for outbound in outbounds:
            if 'voice_speech_url' in outbound.get('metadata', {}):

                data['total'] += 1.0
                if not outbound['delivered']:
                    data['failed'] += 1.0

        if data['failed']:
            data['rate'] = data['failed'] / data['total'] * 100

        sheet.add_row({
            1: "In the last period:",
            2: "{} - {}".format(start_date.strftime('%Y-%m-%d'),
                                end_date.strftime('%Y-%m-%d')),
        })

        sheet.set_header([
            "OBDs Sent",
            "OBDs failed",
            "Failure rate",
        ], row=3)

        sheet.add_row({
            1: data['total'],
            2: data['failed'],
            3: '{0:.2f}%'.format(data.get('rate', 0)),
        })

    def handle_optouts(
            self, sheet, sbm_client, ids_client, start_date, end_date):

        sheet.set_header([
            "MSISDN",
            "Optout Date",
            "Request Source",
            "Reason"
        ])

        optouts = self.get_optouts(
            ids_client, created_at__gte=start_date.isoformat(),
            created_at__lte=end_date.isoformat())

        for optout in optouts:
            msisdns = []
            if 'address' in optout and optout['address'] is not None:
                msisdns = [optout['address']]
            elif 'identity' in optout and optout['identity'] is not None:
                identity = self.get_identity(ids_client, optout['identity'])
                details = identity.get('details', {})
                default_addr_type = details.get('default_addr_type')
                if default_addr_type:
                    addresses = details.get('addresses', {})
                    msisdns = addresses.get(default_addr_type, {}).keys()

            sheet.add_row({
                "MSISDN": ','.join(msisdns),
                "Request Source": optout['request_source'],
                "Reason": optout['reason'],
                "Optout Date": optout['created_at']
            })

generate_report = GenerateReport()

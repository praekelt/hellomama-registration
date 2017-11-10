import pytz
import calendar
import random
import string

from datetime import datetime, timedelta
from django.conf import settings
from openpyxl import Workbook


def midnight(timestamp):
    return timestamp.replace(hour=0, minute=0, second=0, microsecond=0)


def one_month_after(timestamp):
    weekday, number_of_days = calendar.monthrange(
        timestamp.year, timestamp.month)
    return timestamp + timedelta(days=number_of_days)


def midnight_validator(inputstr):
    return midnight(datetime.strptime(inputstr, '%Y-%m-%d')).replace(
        tzinfo=pytz.timezone(settings.TIME_ZONE))


def generate_random_filename(suffix='.xlsx'):
    return ''.join(
        random.choice(string.ascii_lowercase) for i in range(12)) + suffix


class ExportSheet(object):

    def __init__(self, sheet, headers=None):
        self._sheet = sheet
        self.set_header(headers or [])

    def set_header(self, headers, row=1):
        self._headers = headers
        for index, header in enumerate(headers):
            self._sheet.cell(row=row, column=index + 1, value=header)

    def get_header(self):
        return self._headers

    def add_row(self, row):
        row_number = self._sheet.max_row + 1
        for key, value in row.items():
            if isinstance(key, int):
                col_idx = key
            else:
                col_idx = self._headers.index(key) + 1

            cell = self._sheet.cell(
                row=row_number,
                column=col_idx)
            cell.value = value


class ExportWorkbook(object):

    def __init__(self):
        self._workbook = Workbook()

    def add_sheet(self, sheetname, position):
        return ExportSheet(self._workbook.create_sheet(sheetname, position))

    def save(self, file_name):
        return self._workbook.save(file_name)

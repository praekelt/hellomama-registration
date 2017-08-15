from base64 import b64decode
import json
import uuid
from datetime import timedelta, datetime
try:
    from StringIO import StringIO
except ImportError:
    from io import StringIO
import responses
import openpyxl

try:
    from urllib.parse import urlparse
except ImportError:
    from urlparse import urlparse

try:
    import mock
except ImportError:
    from unittest import mock

from django.contrib.auth.models import User
from django.test import TestCase, override_settings
from django.db.models.signals import post_save
from django.conf import settings
from django.core.cache import cache
from django.core.management import call_command
from django.core.management.base import CommandError
from rest_framework import status
from rest_framework.test import APIClient
from rest_framework.authtoken.models import Token
from rest_hooks.models import model_saved, Hook
from requests.exceptions import ConnectTimeout
from requests_testadapter import TestAdapter, TestSession
from openpyxl.writer.excel import save_virtual_workbook

from hellomama_registration import utils
from registrations import tasks
from .models import (
    Source, Registration, SubscriptionRequest, registration_post_save,
    fire_created_metric, fire_unique_operator_metric, fire_message_type_metric,
    fire_source_metric, fire_receiver_type_metric, fire_language_metric,
    fire_state_metric, fire_role_metric, ThirdPartyRegistrationError)
from .tasks import (
    validate_registration,
    is_valid_date, is_valid_uuid, is_valid_lang, is_valid_msg_type,
    is_valid_msg_receiver, is_valid_loss_reason, is_valid_state, is_valid_role,
    repopulate_metrics)


def override_get_today():
    return datetime.strptime("20150817", "%Y%m%d")


class RecordingAdapter(TestAdapter):

    """ Record the request that was handled by the adapter.
    """
    def __init__(self, *args, **kwargs):
        self.requests = []
        super(RecordingAdapter, self).__init__(*args, **kwargs)

    def send(self, request, *args, **kw):
        self.requests.append(request)
        return super(RecordingAdapter, self).send(request, *args, **kw)


REG_FIELDS = {
    "hw_pre_friend": [
        "mother_id", "operator_id", "language", "msg_type",
        "last_period_date", "msg_receiver"]
}

REG_DATA = {
    "hw_pre_mother": {
        "receiver_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
        "operator_id": "nurse000-6a07-4377-a4f6-c0485ccba234",
        "language": "eng_NG",
        "msg_type": "text",
        "gravida": "1",
        "last_period_date": "20150202",
        "msg_receiver": "mother_only"
    },
    "hw_pre_friend": {
        "receiver_id": "friend00-73a2-4d89-b045-d52004c025fe",
        "operator_id": "nurse000-6a07-4377-a4f6-c0485ccba234",
        "language": "eng_NG",
        "msg_type": "text",
        "gravida": "1",
        "last_period_date": "20150202",
        "msg_receiver": "friend_only"
    },
    "hw_pre_family": {
        "receiver_id": "family00-73a2-4d89-b045-d52004c025fe",
        "operator_id": "nurse000-6a07-4377-a4f6-c0485ccba234",
        "language": "eng_NG",
        "msg_type": "text",
        "gravida": "1",
        "last_period_date": "20150202",
        "msg_receiver": "family_only"
    },
    "hw_pre_father": {
        "receiver_id": "father00-73a2-4d89-b045-d52004c025fe",
        "operator_id": "nurse000-6a07-4377-a4f6-c0485ccba234",
        "language": "eng_NG",
        "msg_type": "text",
        "gravida": "2",
        "last_period_date": "20150202",
        "msg_receiver": "father_only"
    },
    "hw_pre_father_and_mother": {
        "receiver_id": "father00-73a2-4d89-b045-d52004c025fe",
        "operator_id": "nurse000-6a07-4377-a4f6-c0485ccba234",
        "language": "eng_NG",
        "msg_type": "text",
        "gravida": "2",
        "last_period_date": "20150202",
        "msg_receiver": "mother_father"
    },
    "hw_pre_family_and_mother": {
        "receiver_id": "family00-73a2-4d89-b045-d52004c025fe",
        "operator_id": "nurse000-6a07-4377-a4f6-c0485ccba234",
        "language": "eng_NG",
        "msg_type": "text",
        "gravida": "2",
        "last_period_date": "20150202",
        "msg_receiver": "mother_family"
    },
    "hw_post": {
        "receiver_id": str(uuid.uuid4()),
        "operator_id": "nurse111-6a07-4377-a4f6-c0485ccba234",
        "language": "eng_NG",
        "msg_type": "text",
        "gravida": "2",
        "baby_dob": "20150202",
        "msg_receiver": "friend_only"
    },
    "pbl_loss": {
        "receiver_id": str(uuid.uuid4()),
        "operator_id": str(uuid.uuid4()),
        "language": "eng_NG",
        "msg_type": "text",
        "gravida": "2",
        "loss_reason": "miscarriage"
    },
    "missing_field": {
        "receiver_id": str(uuid.uuid4()),
        "operator_id": str(uuid.uuid4()),
        "language": "eng_NG",
        "msg_type": "text",
        "gravida": "2",
        "last_period_date": "20150202",
    },
    "bad_fields": {
        "receiver_id": str(uuid.uuid4()),
        "operator_id": str(uuid.uuid4()),
        "language": "eng_NG",
        "msg_type": "text",
        "gravida": "2",
        "last_period_date": "2015020",
        "msg_receiver": "trusted friend"
    },
    "bad_lmp": {
        "receiver_id": str(uuid.uuid4()),
        "operator_id": str(uuid.uuid4()),
        "language": "eng_NG",
        "msg_type": "text",
        "gravida": "2",
        "last_period_date": "20140202",
        "msg_receiver": "friend_only"
    },
}


class APITestCase(TestCase):

    def setUp(self):
        self.adminclient = APIClient()
        self.normalclient = APIClient()
        self.otherclient = APIClient()
        self.session = TestSession()
        utils.get_today = override_get_today


class AuthenticatedAPITestCase(APITestCase):

    def _replace_post_save_hooks(self):
        def has_listeners():
            return post_save.has_listeners(Registration)
        assert has_listeners(), (
            "Registration model has no post_save listeners. Make sure"
            " helpers cleaned up properly in earlier tests.")
        post_save.disconnect(receiver=registration_post_save,
                             sender=Registration)
        post_save.disconnect(receiver=fire_created_metric,
                             sender=Registration)
        post_save.disconnect(receiver=fire_source_metric,
                             sender=Registration)
        post_save.disconnect(receiver=fire_unique_operator_metric,
                             sender=Registration)
        post_save.disconnect(receiver=fire_message_type_metric,
                             sender=Registration)
        post_save.disconnect(receiver=fire_receiver_type_metric,
                             sender=Registration)
        post_save.disconnect(receiver=fire_language_metric,
                             sender=Registration)
        post_save.disconnect(receiver=fire_state_metric,
                             sender=Registration)
        post_save.disconnect(receiver=fire_role_metric,
                             sender=Registration)
        post_save.disconnect(receiver=model_saved,
                             dispatch_uid='instance-saved-hook')
        assert not has_listeners(), (
            "Registration model still has post_save listeners. Make sure"
            " helpers cleaned up properly in earlier tests.")

    def _restore_post_save_hooks(self):
        def has_listeners():
            return post_save.has_listeners(Registration)
        assert not has_listeners(), (
            "Registration model still has post_save listeners. Make sure"
            " helpers removed them properly in earlier tests.")
        post_save.connect(receiver=registration_post_save,
                          sender=Registration)
        post_save.connect(receiver=fire_created_metric,
                          sender=Registration)
        post_save.connect(receiver=fire_source_metric,
                          sender=Registration)
        post_save.connect(receiver=fire_unique_operator_metric,
                          sender=Registration)
        post_save.connect(receiver=fire_language_metric,
                          sender=Registration)
        post_save.connect(receiver=fire_state_metric,
                          sender=Registration)
        post_save.connect(receiver=fire_role_metric,
                          sender=Registration)
        post_save.connect(receiver=model_saved,
                          dispatch_uid='instance-saved-hook')

    def make_source_adminuser(self):
        data = {
            "name": "test_ussd_source_adminuser",
            "authority": "hw_full",
            "user": User.objects.get(username='testadminuser')
        }
        return Source.objects.create(**data)

    def make_source_normaluser(self):
        data = {
            "name": "test_voice_source_normaluser",
            "authority": "patient",
            "user": User.objects.get(username='testnormaluser')
        }
        return Source.objects.create(**data)

    def make_registration_adminuser(self, data=None):
        if data is None:
            data = {
                "stage": "prebirth",
                "data": REG_DATA['hw_pre_mother'],
                "source": self.make_source_adminuser()
            }
        return Registration.objects.create(**data)

    def make_registration_normaluser(self):
        data = {
            "stage": "postbirth",
            "data": REG_DATA['hw_pre_mother'],
            "source": self.make_source_normaluser()
        }
        return Registration.objects.create(**data)

    def setUp(self):
        super(AuthenticatedAPITestCase, self).setUp()
        self._replace_post_save_hooks()

        # Add a user with an email username
        self.emailusername = 'guy@example.com'
        self.emailpassword = 'guypassword'
        self.normaluser = User.objects.create_user(
            self.emailusername,
            'guy@example.com',
            self.emailpassword)

        # Normal User setup
        self.normalusername = 'testnormaluser'
        self.normalpassword = 'testnormalpass'
        self.normaluser = User.objects.create_user(
            self.normalusername,
            'testnormaluser@example.com',
            self.normalpassword)
        normaltoken = Token.objects.create(user=self.normaluser)
        self.normaltoken = normaltoken.key
        self.normalclient.credentials(
            HTTP_AUTHORIZATION='Token ' + self.normaltoken)

        # Admin User setup
        self.adminusername = 'testadminuser'
        self.adminpassword = 'testadminpass'
        self.adminuser = User.objects.create_superuser(
            self.adminusername,
            'testadminuser@example.com',
            self.adminpassword)
        admintoken = Token.objects.create(user=self.adminuser)
        self.admintoken = admintoken.key
        self.adminclient.credentials(
            HTTP_AUTHORIZATION='Token ' + self.admintoken)

    def tearDown(self):
        self._restore_post_save_hooks()


class TestLogin(AuthenticatedAPITestCase):

    def test_login_normaluser(self):
        """ Test that normaluser can login successfully
        """
        # Setup
        post_auth = {"username": "testnormaluser",
                     "password": "testnormalpass"}
        # Execute
        request = self.normalclient.post(
            '/api/token-auth/', post_auth)
        token = request.data.get('token', None)
        # Check
        self.assertIsNotNone(
            token, "Could not receive authentication token on login post.")
        self.assertEqual(
            request.status_code, 200,
            "Status code on /api/token-auth was %s (should be 200)."
            % request.status_code)

    def test_login_adminuser(self):
        """ Test that adminuser can login successfully
        """
        # Setup
        post_auth = {"username": "testadminuser",
                     "password": "testadminpass"}
        # Execute
        request = self.adminclient.post(
            '/api/token-auth/', post_auth)
        token = request.data.get('token', None)
        # Check
        self.assertIsNotNone(
            token, "Could not receive authentication token on login post.")
        self.assertEqual(
            request.status_code, 200,
            "Status code on /api/token-auth was %s (should be 200)."
            % request.status_code)

    def test_login_adminuser_wrong_password(self):
        """ Test that adminuser cannot log in with wrong password
        """
        # Setup
        post_auth = {"username": "testadminuser",
                     "password": "wrongpass"}
        # Execute
        request = self.adminclient.post(
            '/api/token-auth/', post_auth)
        token = request.data.get('token', None)
        # Check
        self.assertIsNone(
            token, "Could not receive authentication token on login post.")
        self.assertEqual(request.status_code, status.HTTP_400_BAD_REQUEST)

    def test_login_otheruser(self):
        """ Test that an unknown user cannot log in
        """
        # Setup
        post_auth = {"username": "testotheruser",
                     "password": "testotherpass"}
        # Execute
        request = self.otherclient.post(
            '/api/token-auth/', post_auth)
        token = request.data.get('token', None)
        # Check
        self.assertIsNone(
            token, "Could not receive authentication token on login post.")
        self.assertEqual(request.status_code, status.HTTP_400_BAD_REQUEST)


class TestSourceAPI(AuthenticatedAPITestCase):

    def test_get_source_adminuser(self):
        # Setup
        source = self.make_source_adminuser()
        # Execute
        response = self.adminclient.get('/api/v1/source/%s/' % source.id,
                                        format='json',
                                        content_type='application/json')
        # Check
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["authority"], "hw_full")
        self.assertEqual(response.data["name"], 'test_ussd_source_adminuser')

    def test_get_source_normaluser(self):
        # Setup
        source = self.make_source_normaluser()
        # Execute
        response = self.normalclient.get('/api/v1/source/%s/' % source.id,
                                         content_type='application/json')
        # Check
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_create_source_adminuser(self):
        # Setup
        user = User.objects.get(username='testadminuser')
        post_data = {
            "name": "test_source_name",
            "authority": "patient",
            "user": "/api/v1/user/%s/" % user.id
        }
        # Execute
        response = self.adminclient.post('/api/v1/source/',
                                         json.dumps(post_data),
                                         content_type='application/json')
        # Check
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        d = Source.objects.last()
        self.assertEqual(d.name, 'test_source_name')
        self.assertEqual(d.authority, "patient")

    def test_create_source_normaluser(self):
        # Setup
        user = User.objects.get(username='testnormaluser')
        post_data = {
            "name": "test_source_name",
            "authority": "hw_full",
            "user": "/api/v1/user/%s/" % user.id
        }
        # Execute
        response = self.normalclient.post('/api/v1/source/',
                                          json.dumps(post_data),
                                          content_type='application/json')
        # Check
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)


class TestRegistrationAPI(AuthenticatedAPITestCase):

    def test_get_registration_adminuser(self):
        # Setup
        registration = self.make_registration_adminuser()
        # Execute
        response = self.adminclient.get(
            '/api/v1/registration/%s/' % registration.id,
            content_type='application/json')
        # Check
        # Currently only posts are allowed
        self.assertEqual(response.status_code,
                         status.HTTP_405_METHOD_NOT_ALLOWED)

    def test_get_registration_normaluser(self):
        # Setup
        registration = self.make_registration_normaluser()
        # Execute
        response = self.normalclient.get(
            '/api/v1/registration/%s/' % registration.id,
            content_type='application/json')
        # Check
        # Currently only posts are allowed
        self.assertEqual(response.status_code,
                         status.HTTP_405_METHOD_NOT_ALLOWED)

    def test_create_registration_adminuser(self):
        # Setup
        self.make_source_adminuser()
        post_data = {
            "stage": "prebirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": {"test_key1": "test_value1"}
        }
        # Execute
        response = self.adminclient.post('/api/v1/registration/',
                                         json.dumps(post_data),
                                         content_type='application/json')
        # Check
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        d = Registration.objects.last()
        self.assertEqual(d.source.name, 'test_ussd_source_adminuser')
        self.assertEqual(d.stage, 'prebirth')
        self.assertEqual(d.validated, False)
        self.assertEqual(d.data, {"test_key1": "test_value1"})
        self.assertEqual(d.created_by, self.adminuser)

    def test_create_registration_normaluser(self):
        # Setup
        self.make_source_normaluser()
        post_data = {
            "stage": "postbirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": {"test_key1": "test_value1"}
        }
        # Execute
        response = self.normalclient.post('/api/v1/registration/',
                                          json.dumps(post_data),
                                          content_type='application/json')
        # Check
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        d = Registration.objects.last()
        self.assertEqual(d.source.name, 'test_voice_source_normaluser')
        self.assertEqual(d.stage, 'postbirth')
        self.assertEqual(d.validated, False)
        self.assertEqual(d.data, {"test_key1": "test_value1"})

    def test_create_registration_set_readonly_field(self):
        # Setup
        self.make_source_adminuser()
        post_data = {
            "stage": "prebirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": {"test_key1": "test_value1"},
            "validated": True
        }
        # Execute
        response = self.adminclient.post('/api/v1/registration/',
                                         json.dumps(post_data),
                                         content_type='application/json')
        # Check
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        d = Registration.objects.last()
        self.assertEqual(d.source.name, 'test_ussd_source_adminuser')
        self.assertEqual(d.stage, 'prebirth')
        self.assertEqual(d.validated, False)  # Should ignore True post_data
        self.assertEqual(d.data, {"test_key1": "test_value1"})

    def test_update_registration_adminuser(self):
        # Setup
        registration = self.make_registration_normaluser()
        post_data = {
            "data": {"test_key1": "test_value1"}
        }
        # Execute
        response = self.adminclient.patch(
            '/api/v1/registration/%s/' % registration.id,
            json.dumps(post_data),
            content_type='application/json')
        # Check
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        registration.refresh_from_db()

        self.assertEqual(registration.data, {"test_key1": "test_value1"})
        self.assertEqual(registration.updated_by, self.adminuser)

    def test_update_registration_normaluser(self):
        # Setup
        registration = self.make_registration_normaluser()
        post_data = {
            "data": {"test_key1": "test_value1"}
        }
        # Execute
        response = self.normalclient.patch(
            '/api/v1/registration/%s/' % registration.id,
            json.dumps(post_data),
            content_type='application/json')
        # Check
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        registration.refresh_from_db()

        self.assertEqual(registration.data, {"test_key1": "test_value1"})
        self.assertEqual(registration.updated_by, self.normaluser)

    def test_update_registration_readonly_field(self):
        # Setup
        registration = self.make_registration_normaluser()
        post_data = {
            "data": {"test_key1": "test_value1"},
            "validated": True
        }
        # Execute
        response = self.adminclient.patch(
            '/api/v1/registration/%s/' % registration.id,
            json.dumps(post_data),
            content_type='application/json')
        # Check
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        registration.refresh_from_db()

        self.assertEqual(registration.data, {"test_key1": "test_value1"})
        self.assertEqual(registration.updated_by, self.adminuser)
        self.assertEqual(registration.validated, False)

    def test_list_registrations(self):
        # Setup
        registration1 = self.make_registration_normaluser()
        registration2 = self.make_registration_adminuser()
        # Execute
        response = self.normalclient.get(
            '/api/v1/registrations/', content_type='application/json')
        # Check
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 2)
        result1, result2 = response.data["results"]
        self.assertEqual(result1["id"], str(registration1.id))
        self.assertEqual(result2["id"], str(registration2.id))

    def make_different_registrations(self):
        self.make_source_adminuser()
        registration1_data = {
            "stage": "prebirth",
            "mother_id": "mother01-63e2-4acc-9b94-26663b9bc267",
            "data": REG_DATA["hw_pre_mother"].copy(),
            "source": self.make_source_adminuser(),
            "validated": True
        }
        registration1 = Registration.objects.create(**registration1_data)
        registration2_data = {
            "stage": "postbirth",
            "mother_id": "mother02-63e2-4acc-9b94-26663b9bc267",
            "data": REG_DATA["hw_pre_friend"].copy(),
            "source": self.make_source_normaluser(),
            "validated": False
        }
        registration2 = Registration.objects.create(**registration2_data)

        return (registration1, registration2)

    def test_filter_registration_mother_id(self):
        # Setup
        registration1, registration2 = self.make_different_registrations()
        # Execute
        response = self.adminclient.get(
            '/api/v1/registrations/?mother_id=%s' % registration1.mother_id,
            content_type='application/json')
        # Check
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 1)
        result = response.data["results"][0]
        self.assertEqual(result["id"], str(registration1.id))

    def test_filter_registration_stage(self):
        # Setup
        registration1, registration2 = self.make_different_registrations()
        # Execute
        response = self.adminclient.get(
            '/api/v1/registrations/?stage=%s' % registration2.stage,
            content_type='application/json')
        # Check
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 1)
        result = response.data["results"][0]
        self.assertEqual(result["id"], str(registration2.id))

    def test_filter_registration_validated(self):
        # Setup
        registration1, registration2 = self.make_different_registrations()
        # Execute
        response = self.adminclient.get(
            '/api/v1/registrations/?validated=%s' % registration1.validated,
            content_type='application/json')
        # Check
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 1)
        result = response.data["results"][0]
        self.assertEqual(result["id"], str(registration1.id))

    def test_filter_registration_source(self):
        # Setup
        registration1, registration2 = self.make_different_registrations()
        # Execute
        response = self.adminclient.get(
            '/api/v1/registrations/?source=%s' % registration2.source.id,
            content_type='application/json')
        # Check
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 1)
        result = response.data["results"][0]
        self.assertEqual(result["id"], str(registration2.id))

    def test_filter_registration_created_after(self):
        # Setup
        registration1, registration2 = self.make_different_registrations()
        # While the '+00:00' is valid according to ISO 8601, the version of
        # django-filter we are using does not support it
        date_string = registration2.created_at.isoformat().replace(
            "+00:00", "Z")
        # Execute
        response = self.adminclient.get(
            '/api/v1/registrations/?created_after=%s' % date_string,
            content_type='application/json')
        # Check
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 1)
        result = response.data["results"][0]
        self.assertEqual(result["id"], str(registration2.id))

    def test_filter_registration_created_before(self):
        # Setup
        registration1, registration2 = self.make_different_registrations()
        # While the '+00:00' is valid according to ISO 8601, the version of
        # django-filter we are using does not support it
        date_string = registration1.created_at.isoformat().replace(
            "+00:00", "Z")
        # Execute
        response = self.adminclient.get(
            '/api/v1/registrations/?created_before=%s' % date_string,
            content_type='application/json')
        # Check
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 1)
        result = response.data["results"][0]
        self.assertEqual(result["id"], str(registration1.id))

    def test_filter_registration_no_matches(self):
        # Setup
        registration1, registration2 = self.make_different_registrations()
        # Execute
        response = self.adminclient.get(
            '/api/v1/registrations/?mother_id=test_id',
            content_type='application/json')
        # Check
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 0)

    def test_filter_registration_unknown_filter(self):
        # Setup
        registration1, registration2 = self.make_different_registrations()
        # Execute
        response = self.adminclient.get(
            '/api/v1/registrations/?something=test_id',
            content_type='application/json')
        # Check
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 2)


class TestFieldValidation(AuthenticatedAPITestCase):

    def test_is_valid_date(self):
        # Setup
        good_date = "19820315"
        invalid_date = "19830229"
        bad_date = "1234"
        # Execute
        # Check
        self.assertEqual(is_valid_date(good_date), True)
        self.assertEqual(is_valid_date(invalid_date), False)
        self.assertEqual(is_valid_date(bad_date), False)

    def test_is_valid_uuid(self):
        # Setup
        valid_uuid = str(uuid.uuid4())
        invalid_uuid = "f9bfa2d7-5b62-4011-8eac-76bca34781a"
        # Execute
        # Check
        self.assertEqual(is_valid_uuid(valid_uuid), True)
        self.assertEqual(is_valid_uuid(invalid_uuid), False)

    def test_is_valid_lang(self):
        # Setup
        valid_lang = "pcm_NG"
        invalid_lang = "pidgin"
        # Execute
        # Check
        self.assertEqual(is_valid_lang(valid_lang), True)
        self.assertEqual(is_valid_lang(invalid_lang), False)

    def test_is_valid_state(self):
        # Setup
        valid_state = "cross_river"
        invalid_state = "new_jersey"
        # Execute
        # Check
        self.assertEqual(is_valid_state(valid_state), True)
        self.assertEqual(is_valid_state(invalid_state), False)

    def test_is_valid_role(self):
        # Setup
        valid_role = "midwife"
        invalid_role = "nurse"
        # Execute
        # Check
        self.assertEqual(is_valid_role(valid_role), True)
        self.assertEqual(is_valid_role(invalid_role), False)

    def test_is_valid_msg_type(self):
        # Setup
        valid_msg_type1 = "text"
        valid_msg_type2 = "audio"
        invalid_msg_type = "email"
        # Execute
        # Check
        self.assertEqual(is_valid_msg_type(valid_msg_type1), True)
        self.assertEqual(is_valid_msg_type(valid_msg_type2), True)
        self.assertEqual(is_valid_msg_type(invalid_msg_type), False)

    def test_is_valid_msg_receiver(self):
        # Setup
        valid_msg_receiver = "father_only"
        invalid_msg_receiver = "mama"
        # Execute
        # Check
        self.assertEqual(is_valid_msg_receiver(valid_msg_receiver), True)
        self.assertEqual(is_valid_msg_receiver(invalid_msg_receiver), False)

    def test_is_valid_loss_reason(self):
        # Setup
        valid_loss_reason = "miscarriage"
        invalid_loss_reason = "other"
        # Execute
        # Check
        self.assertEqual(is_valid_loss_reason(valid_loss_reason), True)
        self.assertEqual(is_valid_loss_reason(invalid_loss_reason), False)

    def test_check_field_values(self):
        # Setup
        valid_hw_pre_registration_data = REG_DATA["hw_pre_friend"].copy()
        invalid_hw_pre_registration_data = REG_DATA["hw_pre_friend"].copy()
        invalid_hw_pre_registration_data["msg_receiver"] = "somebody"
        # Execute
        cfv_valid = validate_registration.check_field_values(
            REG_FIELDS["hw_pre_friend"], valid_hw_pre_registration_data)
        cfv_invalid = validate_registration.check_field_values(
            REG_FIELDS["hw_pre_friend"], invalid_hw_pre_registration_data)
        # Check
        self.assertEqual(cfv_valid, [])
        self.assertEqual(cfv_invalid, ['msg_receiver'])


class TestRegistrationValidation(AuthenticatedAPITestCase):

    def test_validate_hw_prebirth(self):
        # Setup
        registration_data = {
            "stage": "prebirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": REG_DATA["hw_pre_friend"].copy(),
            "source": self.make_source_adminuser()
        }
        registration = Registration.objects.create(**registration_data)
        # Execute
        v = validate_registration.validate(registration)
        # Check
        self.assertEqual(v, True)
        self.assertEqual(registration.data["reg_type"], "hw_pre")
        self.assertEqual(registration.data["preg_week"], 28)
        self.assertEqual(registration.validated, True)

    def test_validate_hw_postbirth(self):
        # Setup
        registration_data = {
            "stage": "postbirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": REG_DATA["hw_post"].copy(),
            "source": self.make_source_adminuser()
        }
        registration = Registration.objects.create(**registration_data)
        # Execute
        v = validate_registration.validate(registration)
        # Check
        self.assertEqual(v, True)
        self.assertEqual(registration.data["reg_type"], "hw_post")
        self.assertEqual(registration.data["baby_age"], 28)
        self.assertEqual(registration.validated, True)

    def test_validate_pbl_loss(self):
        # Setup
        registration_data = {
            "stage": "loss",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": REG_DATA["pbl_loss"].copy(),
            "source": self.make_source_normaluser()
        }
        registration = Registration.objects.create(**registration_data)
        # Execute
        v = validate_registration.validate(registration)
        # Check
        self.assertEqual(v, True)
        self.assertEqual(registration.data["reg_type"], "pbl_loss")
        self.assertEqual(registration.validated, True)

    def test_validate_pregnancy_too_long(self):
        # Setup
        registration_data = {
            "stage": "prebirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": REG_DATA["hw_pre_friend"].copy(),
            "source": self.make_source_adminuser()
        }
        registration_data["data"]["last_period_date"] = "20130101"
        registration = Registration.objects.create(**registration_data)
        # Execute
        v = validate_registration.validate(registration)
        # Check
        self.assertEqual(v, False)
        self.assertEqual(registration.validated, False)

    def test_validate_pregnancy_9_weeks(self):
        # Setup
        registration_data = {
            "stage": "prebirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": REG_DATA["hw_pre_friend"].copy(),
            "source": self.make_source_adminuser()
        }
        registration_data["data"]["last_period_date"] = "20150612"  # 9 weeks
        registration = Registration.objects.create(**registration_data)
        # Execute
        v = validate_registration.validate(registration)
        # Check
        self.assertEqual(v, False)
        self.assertEqual(registration.validated, False)

    def test_validate_pregnancy_10_weeks(self):
        # Setup
        registration_data = {
            "stage": "prebirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": REG_DATA["hw_pre_friend"].copy(),
            "source": self.make_source_adminuser()
        }
        registration_data["data"]["last_period_date"] = "20150605"  # 10 weeks
        registration = Registration.objects.create(**registration_data)
        # Execute
        v = validate_registration.validate(registration)
        # Check
        self.assertEqual(v, True)
        self.assertEqual(registration.validated, True)

    def test_validate_baby_too_young(self):
        # Setup
        registration_data = {
            "stage": "postbirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": REG_DATA["hw_post"].copy(),
            "source": self.make_source_adminuser()
        }
        registration_data["data"]["baby_dob"] = "20150818"
        registration = Registration.objects.create(**registration_data)
        # Execute
        v = validate_registration.validate(registration)
        # Check
        self.assertEqual(v, False)
        self.assertEqual(registration.validated, False)

    def test_validate_baby_too_old(self):
        # Setup
        registration_data = {
            "stage": "postbirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": REG_DATA["hw_post"].copy(),
            "source": self.make_source_adminuser()
        }
        registration_data["data"]["baby_dob"] = "20130717"
        registration = Registration.objects.create(**registration_data)
        # Execute
        v = validate_registration.validate(registration)
        # Check
        self.assertEqual(v, False)
        self.assertEqual(registration.validated, False)

    @responses.activate
    def test_validate_registration_run_success(self):
        # Setup
        # mock mother messageset lookup
        query_string = '?short_name=prebirth.mother.text.10_42'
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/messageset/%s' % query_string,
            json={
                "count": 1,
                "next": None,
                "previous": None,
                "results": [{
                    "id": 1,
                    "short_name": 'prebirth.mother.text.10_42',
                    "default_schedule": 1
                }]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )
        # mock household messageset lookup
        query_string = '?short_name=prebirth.household.audio.10_42.fri.9_11'
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/messageset/%s' % query_string,
            json={
                "count": 1,
                "next": None,
                "previous": None,
                "results": [{
                    "id": 3,
                    "short_name": 'prebirth.household.audio.10_42.fri.9_11',
                    "default_schedule": 3
                }]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )
        # mock mother schedule lookup
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/schedule/1/',
            json={"id": 1, "day_of_week": "1,3,5"},
            status=200, content_type='application/json',
        )
        # mock household schedule lookup
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/schedule/3/',
            json={"id": 3, "day_of_week": "5"},
            status=200, content_type='application/json',
        )
        # mock mother MSISDN lookup
        responses.add(
            responses.GET,
            'http://localhost:8001/api/v1/identities/mother00-9d89-4aa6-99ff-13c225365b5d/addresses/msisdn?default=True',  # noqa
            json={
                "count": 1, "next": None, "previous": None,
                "results": [{"address": "+234123"}]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )
        # mock friend MSISDN lookup
        responses.add(
            responses.GET,
            'http://localhost:8001/api/v1/identities/friend00-73a2-4d89-b045-d52004c025fe/addresses/msisdn?default=True',  # noqa
            json={
                "count": 1, "next": None, "previous": None,
                "results": [{"address": "+234124"}]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )

        # mock mother welcome SMS send
        responses.add(
            responses.POST,
            'http://localhost:8006/api/v1/outbound/',
            json={"id": 1},
            status=200, content_type='application/json',
        )
        # prepare registration data
        registration_data = {
            "stage": "prebirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": REG_DATA["hw_pre_friend"].copy(),
            "source": self.make_source_adminuser()
        }
        registration = Registration.objects.create(**registration_data)
        # Execute
        result = validate_registration.apply_async(args=[registration.id])
        # Check
        self.assertEqual(result.get(), "Validation completed - Success")

    def test_validate_registration_run_failure_missing_field(self):
        # Setup
        registration_data = {
            "stage": "prebirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": REG_DATA["missing_field"].copy(),
            "source": self.make_source_adminuser()
        }
        registration = Registration.objects.create(**registration_data)
        # Execute
        result = validate_registration.apply_async(args=[registration.id])
        # Check
        self.assertEqual(result.get(), "Validation completed - Failure")
        d = Registration.objects.get(id=registration.id)
        self.assertEqual(d.data["invalid_fields"],
                         "Invalid combination of fields")

    def test_validate_registration_run_failure_bad_fields(self):
        # Setup
        registration_data = {
            "stage": "prebirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": REG_DATA["bad_fields"].copy(),
            "source": self.make_source_adminuser()
        }
        registration = Registration.objects.create(**registration_data)
        # Execute
        result = validate_registration.apply_async(args=[registration.id])
        # Check
        self.assertEqual(result.get(), "Validation completed - Failure")
        d = Registration.objects.get(id=registration.id)
        self.assertEqual(sorted(d.data["invalid_fields"]),
                         sorted(["msg_receiver", "last_period_date"]))

    def test_validate_registration_run_failure_bad_lmp(self):
        # Setup
        registration_data = {
            "stage": "prebirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": REG_DATA["bad_lmp"].copy(),
            "source": self.make_source_adminuser()
        }
        registration = Registration.objects.create(**registration_data)
        # Execute
        result = validate_registration.apply_async(args=[registration.id])
        # Check
        self.assertEqual(result.get(), "Validation completed - Failure")
        d = Registration.objects.get(id=registration.id)
        self.assertEqual(d.data["invalid_fields"],
                         ["last_period_date out of range"])

    def test_validate_registration_run_failure_receiver_id(self):
        # Setup
        registration_data = {
            "stage": "prebirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": REG_DATA["hw_pre_friend"].copy(),
            "source": self.make_source_adminuser()
        }
        # reg_data = registration_data.copy()
        registration_data["data"]["receiver_id"] = registration_data[
            "mother_id"]
        registration = Registration.objects.create(**registration_data)
        # Execute
        result = validate_registration.apply_async(args=[registration.id])
        # Check
        self.assertEqual(result.get(), "Validation completed - Failure")
        d = Registration.objects.get(id=registration.id)
        self.assertEqual(d.data["invalid_fields"], "mother requires own id")

    def test_validate_registration_run_failure_mother_uuid(self):
        # Setup
        registration_data = {
            "stage": "prebirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5",
            "data": REG_DATA["hw_pre_mother"].copy(),
            "source": self.make_source_adminuser()
        }
        registration = Registration.objects.create(**registration_data)
        # Execute
        result = validate_registration.apply_async(args=[registration.id])
        # Check
        self.assertEqual(result.get(), "Validation completed - Failure")
        d = Registration.objects.get(id=registration.id)
        self.assertEqual(d.data["invalid_fields"], "Invalid UUID mother_id")

    def test_validate_registration_run_failure_mother_id(self):
        # Setup
        registration_data = {
            "stage": "prebirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": REG_DATA["hw_pre_mother"].copy(),
            "source": self.make_source_adminuser()
        }
        # reg_data = registration_data.copy()
        registration_data["data"]["receiver_id"] = str(uuid.uuid4())
        registration = Registration.objects.create(**registration_data)
        # Execute
        result = validate_registration.apply_async(args=[registration.id])
        # Check
        self.assertEqual(result.get(), "Validation completed - Failure")
        d = Registration.objects.get(id=registration.id)
        self.assertEqual(d.data["invalid_fields"],
                         "mother_id should be the same as receiver_id")


class TestSubscriptionRequest(AuthenticatedAPITestCase):

    @responses.activate
    def test_mother_only_prebirth_sms(self):
        # Setup
        # mock mother messageset lookup
        query_string = '?short_name=prebirth.mother.text.10_42'
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/messageset/%s' % query_string,
            json={
                "count": 1,
                "next": None,
                "previous": None,
                "results": [{
                    "id": 1,
                    "short_name": 'prebirth.mother.text.10_42',
                    "default_schedule": 1
                }]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )
        # mock mother schedule lookup
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/schedule/1/',
            json={"id": 1, "day_of_week": "1,3,5"},
            status=200, content_type='application/json',
        )
        # mock household schedule lookup
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/schedule/3/',
            json={"id": 3, "day_of_week": "5"},
            status=200, content_type='application/json',
        )
        # mock mother MSISDN lookup
        responses.add(
            responses.GET,
            'http://localhost:8001/api/v1/identities/mother00-9d89-4aa6-99ff-13c225365b5d/addresses/msisdn?default=True',  # noqa
            json={
                "count": 1, "next": None, "previous": None,
                "results": [{"address": "+234123"}]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )

        # mock mother SMS send
        responses.add(
            responses.POST,
            'http://localhost:8006/api/v1/outbound/',
            json={"id": 1},
            status=200, content_type='application/json',
        )

        # prepare registration data
        registration_data = {
            "stage": "prebirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": REG_DATA["hw_pre_mother"].copy(),
            "source": self.make_source_adminuser()
        }
        registration_data["data"]["preg_week"] = 15
        registration = Registration.objects.create(**registration_data)
        # Execute
        result = validate_registration.create_subscriptionrequests(
            registration)
        # Check
        self.assertEqual(result, "1 SubscriptionRequest created")
        d_mom = SubscriptionRequest.objects.last()
        self.assertEqual(d_mom.identity,
                         "mother00-9d89-4aa6-99ff-13c225365b5d")
        self.assertEqual(d_mom.messageset, 1)
        self.assertEqual(d_mom.next_sequence_number, 15)
        self.assertEqual(d_mom.lang, "eng_NG")
        self.assertEqual(d_mom.schedule, 1)

    @responses.activate
    def test_mother_only_prebirth_voice_tue_thu_9_11(self):
        # Setup
        # mock mother messageset lookup
        query_string = '?short_name=prebirth.mother.audio.10_42.tue_thu.9_11'
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/messageset/%s' % query_string,
            json={
                "count": 1,
                "next": None,
                "previous": None,
                "results": [{
                    "id": 2,
                    "short_name": 'prebirth.mother.audio.10_42.tue_thu.9_11',
                    "default_schedule": 6
                }]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )
        # mock mother schedule lookup
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/schedule/6/',
            json={"id": 6, "day_of_week": "2,4"},
            status=200, content_type='application/json',
            match_querystring=True
        )
        # prepare registration data
        registration_data = {
            "stage": "prebirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": REG_DATA["hw_pre_mother"].copy(),
            "source": self.make_source_adminuser()
        }
        registration_data["data"]["preg_week"] = 15
        registration_data["data"]["msg_type"] = "audio"
        registration_data["data"]["voice_times"] = "9_11"
        registration_data["data"]["voice_days"] = "tue_thu"
        registration = Registration.objects.create(**registration_data)
        # Execute
        result = validate_registration.create_subscriptionrequests(
            registration)
        # Check
        self.assertEqual(result, "1 SubscriptionRequest created")
        d_mom = SubscriptionRequest.objects.last()
        self.assertEqual(d_mom.identity,
                         "mother00-9d89-4aa6-99ff-13c225365b5d")
        self.assertEqual(d_mom.messageset, 2)
        self.assertEqual(d_mom.next_sequence_number, 10)
        self.assertEqual(d_mom.lang, "eng_NG")
        self.assertEqual(d_mom.schedule, 6)

    @responses.activate
    def test_friend_only_prebirth_sms(self):
        # Setup
        # mock mother messageset lookup
        query_string = '?short_name=prebirth.mother.text.10_42'
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/messageset/%s' % query_string,
            json={
                "count": 1,
                "next": None,
                "previous": None,
                "results": [{
                    "id": 1,
                    "short_name": 'prebirth.mother.text.10_42',
                    "default_schedule": 1
                }]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )
        # mock household messageset lookup
        query_string = '?short_name=prebirth.household.audio.10_42.fri.9_11'
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/messageset/%s' % query_string,
            json={
                "count": 1,
                "next": None,
                "previous": None,
                "results": [{
                    "id": 3,
                    "short_name": 'prebirth.household.audio.10_42.fri.9_11',
                    "default_schedule": 3
                }]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )
        # mock mother schedule lookup
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/schedule/1/',
            json={"id": 1, "day_of_week": "1,3,5"},
            status=200, content_type='application/json',
        )
        # mock household schedule lookup
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/schedule/3/',
            json={"id": 3, "day_of_week": "5"},
            status=200, content_type='application/json',
        )
        # mock mother MSISDN lookup
        responses.add(
            responses.GET,
            'http://localhost:8001/api/v1/identities/mother00-9d89-4aa6-99ff-13c225365b5d/addresses/msisdn?default=True',  # noqa
            json={
                "count": 1, "next": None, "previous": None,
                "results": [{"address": "+234123"}]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )

        # mock friend MSISDN lookup
        responses.add(
            responses.GET,
            'http://localhost:8001/api/v1/identities/friend00-73a2-4d89-b045-d52004c025fe/addresses/msisdn?default=True',  # noqa
            json={
                "count": 1, "next": None, "previous": None,
                "results": [{"address": "+234123"}]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )

        # mock mother SMS send
        responses.add(
            responses.POST,
            'http://localhost:8006/api/v1/outbound/',
            json={"id": 1},
            status=200, content_type='application/json',
        )
        # prepare registration data
        registration_data = {
            "stage": "prebirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": REG_DATA["hw_pre_friend"].copy(),
            "source": self.make_source_adminuser()
        }
        registration_data["data"]["preg_week"] = 15
        registration = Registration.objects.create(**registration_data)
        # Execute
        result = validate_registration.create_subscriptionrequests(
            registration)
        # Check
        self.assertEqual(result, "2 SubscriptionRequests created")

        d_mom = SubscriptionRequest.objects.get(
            identity="mother00-9d89-4aa6-99ff-13c225365b5d")
        self.assertEqual(d_mom.identity,
                         "mother00-9d89-4aa6-99ff-13c225365b5d")
        self.assertEqual(d_mom.messageset, 1)
        self.assertEqual(d_mom.next_sequence_number, 15)
        self.assertEqual(d_mom.lang, "eng_NG")
        self.assertEqual(d_mom.schedule, 1)

        d_friend = SubscriptionRequest.objects.get(
            identity="friend00-73a2-4d89-b045-d52004c025fe")
        self.assertEqual(d_friend.identity,
                         "friend00-73a2-4d89-b045-d52004c025fe")
        self.assertEqual(d_friend.messageset, 3)
        self.assertEqual(d_friend.next_sequence_number, 5)
        self.assertEqual(d_friend.lang, "eng_NG")
        self.assertEqual(d_friend.schedule, 3)

    @responses.activate
    def test_friend_only_voice_mon_wed_2_5(self):
        # Setup
        # mock mother messageset lookup
        query_string = '?short_name=prebirth.mother.audio.10_42.mon_wed.2_5'
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/messageset/%s' % query_string,
            json={
                "count": 1,
                "next": None,
                "previous": None,
                "results": [{
                    "id": 2,
                    "short_name": 'prebirth.mother.audio.10_42.mon_wed.2_5',
                    "default_schedule": 5
                }]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )
        # mock household messageset lookup
        query_string = '?short_name=prebirth.household.audio.10_42.fri.9_11'
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/messageset/%s' % query_string,
            json={
                "count": 1,
                "next": None,
                "previous": None,
                "results": [{
                    "id": 3,
                    "short_name": 'prebirth.household.audio.10_42.fri.9_11',
                    "default_schedule": 3
                }]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )
        # mock mother schedule lookup
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/schedule/5/',
            json={"id": 5, "day_of_week": "1,3"},
            status=200, content_type='application/json',
            match_querystring=True
        )
        # mock household schedule lookup
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/schedule/3/',
            json={"id": 3, "day_of_week": "5"},
            status=200, content_type='application/json',
        )
        # prepare registration data
        registration_data = {
            "stage": "prebirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": REG_DATA["hw_pre_friend"].copy(),
            "source": self.make_source_adminuser()
        }
        registration_data["data"]["preg_week"] = 15
        registration_data["data"]["msg_type"] = "audio"
        registration_data["data"]["voice_times"] = "2_5"
        registration_data["data"]["voice_days"] = "mon_wed"
        registration = Registration.objects.create(**registration_data)

        # Execute
        result = validate_registration.create_subscriptionrequests(
            registration)

        # Check
        self.assertEqual(result, "2 SubscriptionRequests created")

        d_mom = SubscriptionRequest.objects.get(
            identity="mother00-9d89-4aa6-99ff-13c225365b5d")
        self.assertEqual(d_mom.identity,
                         "mother00-9d89-4aa6-99ff-13c225365b5d")
        self.assertEqual(d_mom.messageset, 2)
        self.assertEqual(d_mom.next_sequence_number, 10)
        self.assertEqual(d_mom.lang, "eng_NG")
        self.assertEqual(d_mom.schedule, 5)
        self.assertEqual(d_mom.metadata["prepend_next_delivery"],
                         "http://registration.dev.example.org/static/audio/registration/eng_NG/welcome_mother.mp3")  # noqa

        d_friend = SubscriptionRequest.objects.get(
            identity="friend00-73a2-4d89-b045-d52004c025fe")
        self.assertEqual(d_friend.identity,
                         "friend00-73a2-4d89-b045-d52004c025fe")
        self.assertEqual(d_friend.messageset, 3)
        self.assertEqual(d_friend.next_sequence_number, 5)
        self.assertEqual(d_friend.lang, "eng_NG")
        self.assertEqual(d_friend.schedule, 3)
        self.assertEqual(d_friend.metadata["prepend_next_delivery"],
                         "http://registration.dev.example.org/static/audio/registration/eng_NG/welcome_household.mp3")  # noqa

    @responses.activate
    def test_family_only_prebirth_sms(self):
        # Setup
        # mock mother messageset lookup
        query_string = '?short_name=prebirth.mother.text.10_42'
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/messageset/%s' % query_string,
            json={
                "count": 1,
                "next": None,
                "previous": None,
                "results": [{
                    "id": 1,
                    "short_name": 'prebirth.mother.text.10_42',
                    "default_schedule": 1
                }]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )
        # mock household messageset lookup
        query_string = '?short_name=prebirth.household.audio.10_42.fri.9_11'
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/messageset/%s' % query_string,
            json={
                "count": 1,
                "next": None,
                "previous": None,
                "results": [{
                    "id": 3,
                    "short_name": 'prebirth.household.audio.10_42.fri.9_11',
                    "default_schedule": 3
                }]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )
        # mock mother schedule lookup
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/schedule/1/',
            json={"id": 1, "day_of_week": "1,3,5"},
            status=200, content_type='application/json',
        )
        # mock household schedule lookup
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/schedule/3/',
            json={"id": 3, "day_of_week": "5"},
            status=200, content_type='application/json',
        )

        # mock mother MSISDN lookup
        responses.add(
            responses.GET,
            'http://localhost:8001/api/v1/identities/mother00-9d89-4aa6-99ff-13c225365b5d/addresses/msisdn?default=True',  # noqa
            json={
                "count": 1, "next": None, "previous": None,
                "results": [{"address": "+234123"}]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )

        # mock family MSISDN lookup
        responses.add(
            responses.GET,
            'http://localhost:8001/api/v1/identities/family00-73a2-4d89-b045-d52004c025fe/addresses/msisdn?default=True',  # noqa
            json={
                "count": 1, "next": None, "previous": None,
                "results": [{"address": "+234124"}]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )

        # mock mother SMS send
        responses.add(
            responses.POST,
            'http://localhost:8006/api/v1/outbound/',
            json={"id": 1},
            status=200, content_type='application/json',
        )

        # prepare registration data
        registration_data = {
            "stage": "prebirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": REG_DATA["hw_pre_family"].copy(),
            "source": self.make_source_adminuser()
        }
        registration_data["data"]["preg_week"] = 15
        registration = Registration.objects.create(**registration_data)
        # Execute
        result = validate_registration.create_subscriptionrequests(
            registration)
        # Check
        self.assertEqual(result, "2 SubscriptionRequests created")

        d_mom = SubscriptionRequest.objects.get(
            identity="mother00-9d89-4aa6-99ff-13c225365b5d")
        self.assertEqual(d_mom.identity,
                         "mother00-9d89-4aa6-99ff-13c225365b5d")
        self.assertEqual(d_mom.messageset, 1)
        self.assertEqual(d_mom.next_sequence_number, 15)
        self.assertEqual(d_mom.lang, "eng_NG")
        self.assertEqual(d_mom.schedule, 1)

        d_family = SubscriptionRequest.objects.get(
            identity="family00-73a2-4d89-b045-d52004c025fe")
        self.assertEqual(d_family.identity,
                         "family00-73a2-4d89-b045-d52004c025fe")
        self.assertEqual(d_family.messageset, 3)
        self.assertEqual(d_family.next_sequence_number, 5)
        self.assertEqual(d_family.lang, "eng_NG")
        self.assertEqual(d_family.schedule, 3)

    @responses.activate
    def test_mother_and_father_prebirth_sms(self):
        # Setup
        # mock mother messageset lookup
        query_string = '?short_name=prebirth.mother.text.10_42'
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/messageset/%s' % query_string,
            json={
                "count": 1,
                "next": None,
                "previous": None,
                "results": [{
                    "id": 1,
                    "short_name": 'prebirth.mother.text.10_42',
                    "default_schedule": 1
                }]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )
        # mock household messageset lookup
        query_string = '?short_name=prebirth.household.audio.10_42.fri.9_11'
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/messageset/%s' % query_string,
            json={
                "count": 1,
                "next": None,
                "previous": None,
                "results": [{
                    "id": 3,
                    "short_name": 'prebirth.household.audio.10_42.fri.9_11',
                    "default_schedule": 3
                }]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )
        # mock mother schedule lookup
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/schedule/1/',
            json={"id": 1, "day_of_week": "1,3,5"},
            status=200, content_type='application/json',
        )
        # mock household schedule lookup
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/schedule/3/',
            json={"id": 3, "day_of_week": "5"},
            status=200, content_type='application/json',
        )
        # mock mother MSISDN lookup
        responses.add(
            responses.GET,
            'http://localhost:8001/api/v1/identities/mother00-9d89-4aa6-99ff-13c225365b5d/addresses/msisdn?default=True',  # noqa
            json={
                "count": 1, "next": None, "previous": None,
                "results": [{"address": "+234123"}]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )
        # mock father MSISDN lookup
        responses.add(
            responses.GET,
            'http://localhost:8001/api/v1/identities/father00-73a2-4d89-b045-d52004c025fe/addresses/msisdn?default=True',  # noqa
            json={
                "count": 1, "next": None, "previous": None,
                "results": [{"address": "+234124"}]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )

        # mock mother SMS send
        responses.add(
            responses.POST,
            'http://localhost:8006/api/v1/outbound/',
            json={"id": 1},
            status=200, content_type='application/json',
        )
        # prepare registration data
        registration_data = {
            "stage": "prebirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": REG_DATA["hw_pre_father_and_mother"].copy(),
            "source": self.make_source_adminuser()
        }
        registration_data["data"]["preg_week"] = 30
        registration = Registration.objects.create(**registration_data)

        # Execute
        result = validate_registration.create_subscriptionrequests(
            registration)

        # Check
        self.assertEqual(result, "2 SubscriptionRequests created")

        d_mom = SubscriptionRequest.objects.get(
            identity="mother00-9d89-4aa6-99ff-13c225365b5d")
        self.assertEqual(d_mom.identity,
                         "mother00-9d89-4aa6-99ff-13c225365b5d")
        self.assertEqual(d_mom.messageset, 1)
        self.assertEqual(d_mom.next_sequence_number, 60)
        self.assertEqual(d_mom.lang, "eng_NG")
        self.assertEqual(d_mom.schedule, 1)

        d_dad = SubscriptionRequest.objects.get(
            identity="father00-73a2-4d89-b045-d52004c025fe")
        self.assertEqual(d_dad.identity,
                         "father00-73a2-4d89-b045-d52004c025fe")
        self.assertEqual(d_dad.messageset, 3)
        self.assertEqual(d_dad.next_sequence_number, 20)
        self.assertEqual(d_dad.lang, "eng_NG")
        self.assertEqual(d_dad.schedule, 3)

    @responses.activate
    def test_mother_and_family_prebirth_sms(self):
        # Setup
        # mock mother messageset lookup
        query_string = '?short_name=prebirth.mother.text.10_42'
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/messageset/%s' % query_string,
            json={
                "count": 1,
                "next": None,
                "previous": None,
                "results": [{
                    "id": 1,
                    "short_name": 'prebirth.mother.text.10_42',
                    "default_schedule": 1
                }]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )
        # mock household messageset lookup
        query_string = '?short_name=prebirth.household.audio.10_42.fri.9_11'
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/messageset/%s' % query_string,
            json={
                "count": 1,
                "next": None,
                "previous": None,
                "results": [{
                    "id": 3,
                    "short_name": 'prebirth.household.audio.10_42.fri.9_11',
                    "default_schedule": 3
                }]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )
        # mock mother schedule lookup
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/schedule/1/',
            json={"id": 1, "day_of_week": "1,3,5"},
            status=200, content_type='application/json',
        )
        # mock household schedule lookup
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/schedule/3/',
            json={"id": 3, "day_of_week": "5"},
            status=200, content_type='application/json',
        )
        # mock mother MSISDN lookup
        responses.add(
            responses.GET,
            'http://localhost:8001/api/v1/identities/mother00-9d89-4aa6-99ff-13c225365b5d/addresses/msisdn?default=True',  # noqa
            json={
                "count": 1, "next": None, "previous": None,
                "results": [{"address": "+234123"}]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )

        # mock family MSISDN lookup
        responses.add(
            responses.GET,
            'http://localhost:8001/api/v1/identities/family00-73a2-4d89-b045-d52004c025fe/addresses/msisdn?default=True',  # noqa
            json={
                "count": 1, "next": None, "previous": None,
                "results": [{"address": "+234124"}]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )

        # mock mother SMS send
        responses.add(
            responses.POST,
            'http://localhost:8006/api/v1/outbound/',
            json={"id": 1},
            status=200, content_type='application/json',
        )
        # prepare registration data
        registration_data = {
            "stage": "prebirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": REG_DATA["hw_pre_family_and_mother"].copy(),
            "source": self.make_source_adminuser()
        }
        registration_data["data"]["preg_week"] = 40
        registration = Registration.objects.create(**registration_data)

        # Execute
        result = validate_registration.create_subscriptionrequests(
            registration)

        # Check
        self.assertEqual(result, "2 SubscriptionRequests created")

        d_mom = SubscriptionRequest.objects.get(
            identity="mother00-9d89-4aa6-99ff-13c225365b5d")
        self.assertEqual(d_mom.identity,
                         "mother00-9d89-4aa6-99ff-13c225365b5d")
        self.assertEqual(d_mom.messageset, 1)
        self.assertEqual(d_mom.next_sequence_number, 90)
        self.assertEqual(d_mom.lang, "eng_NG")
        self.assertEqual(d_mom.schedule, 1)

        d_family = SubscriptionRequest.objects.get(
            identity="family00-73a2-4d89-b045-d52004c025fe")
        self.assertEqual(d_family.identity,
                         "family00-73a2-4d89-b045-d52004c025fe")
        self.assertEqual(d_family.messageset, 3)
        self.assertEqual(d_family.next_sequence_number, 30)
        self.assertEqual(d_family.lang, "eng_NG")
        self.assertEqual(d_family.schedule, 3)


class TestMetricsAPI(AuthenticatedAPITestCase):

    def test_metrics_read(self):
        # Setup
        self.make_source_normaluser()
        self.make_source_adminuser()
        # Execute
        response = self.adminclient.get(
            '/api/metrics/', content_type='application/json')
        # Check
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            sorted(response.data["metrics_available"]), sorted([
                'registrations.created.sum',
                'registrations.created.total.last',
                'registrations.unique_operators.sum',
                'registrations.msg_type.text.sum',
                'registrations.msg_type.audio.sum',
                'registrations.msg_type.text.total.last',
                'registrations.msg_type.audio.total.last',
                'registrations.receiver_type.mother_father.sum',
                'registrations.receiver_type.mother_only.sum',
                'registrations.receiver_type.father_only.sum',
                'registrations.receiver_type.mother_family.sum',
                'registrations.receiver_type.mother_friend.sum',
                'registrations.receiver_type.friend_only.sum',
                'registrations.receiver_type.family_only.sum',
                'registrations.receiver_type.mother_father.total.last',
                'registrations.receiver_type.mother_only.total.last',
                'registrations.receiver_type.father_only.total.last',
                'registrations.receiver_type.mother_family.total.last',
                'registrations.receiver_type.mother_friend.total.last',
                'registrations.receiver_type.friend_only.total.last',
                'registrations.receiver_type.family_only.total.last',
                'registrations.language.eng_NG.sum',
                'registrations.language.hau_NG.sum',
                'registrations.language.ibo_NG.sum',
                'registrations.language.yor_NG.sum',
                'registrations.language.pcm_NG.sum',
                'registrations.language.eng_NG.total.last',
                'registrations.language.hau_NG.total.last',
                'registrations.language.ibo_NG.total.last',
                'registrations.language.yor_NG.total.last',
                'registrations.language.pcm_NG.total.last',
                'registrations.state.ebonyi.sum',
                'registrations.state.cross_river.sum',
                'registrations.state.abuja.sum',
                'registrations.state.ebonyi.total.last',
                'registrations.state.cross_river.total.last',
                'registrations.state.abuja.total.last',
                'registrations.role.oic.sum',
                'registrations.role.cv.sum',
                'registrations.role.midwife.sum',
                'registrations.role.chew.sum',
                'registrations.role.mama.sum',
                'registrations.role.oic.total.last',
                'registrations.role.cv.total.last',
                'registrations.role.midwife.total.last',
                'registrations.role.chew.total.last',
                'registrations.role.mama.total.last',
                'registrations.source.testnormaluser.sum',
                'registrations.source.testadminuser.sum',
                'registrations.change.language.sum',
                'registrations.change.language.total.last',
                'registrations.change.pregnant_to_baby.sum',
                'registrations.change.pregnant_to_baby.total.last',
                'registrations.change.pregnant_to_loss.sum',
                'registrations.change.pregnant_to_loss.total.last',
                'registrations.change.messaging.sum',
                'registrations.change.messaging.total.last',
                'optout.receiver_type.mother_father.sum',
                'optout.receiver_type.mother_only.sum',
                'optout.receiver_type.father_only.sum',
                'optout.receiver_type.mother_family.sum',
                'optout.receiver_type.mother_friend.sum',
                'optout.receiver_type.friend_only.sum',
                'optout.receiver_type.family_only.sum',
                'optout.receiver_type.mother_father.total.last',
                'optout.receiver_type.mother_only.total.last',
                'optout.receiver_type.father_only.total.last',
                'optout.receiver_type.mother_family.total.last',
                'optout.receiver_type.mother_friend.total.last',
                'optout.receiver_type.friend_only.total.last',
                'optout.receiver_type.family_only.total.last',
                'optout.reason.miscarriage.sum',
                'optout.reason.stillborn.sum',
                'optout.reason.baby_death.sum',
                'optout.reason.not_useful.sum',
                'optout.reason.other.sum',
                'optout.reason.miscarriage.total.last',
                'optout.reason.stillborn.total.last',
                'optout.reason.baby_death.total.last',
                'optout.reason.not_useful.total.last',
                'optout.reason.other.total.last',
                'optout.msg_type.audio.sum',
                'optout.msg_type.audio.total.last',
                'optout.msg_type.text.sum',
                'optout.msg_type.text.total.last',
                'optout.source.ussd.sum',
                'optout.source.sms.sum',
                'optout.source.voice.sum',
                'optout.source.ussd.total.last',
                'optout.source.sms.total.last',
                'optout.source.voice.total.last',
            ])
        )

    @responses.activate
    def test_post_metrics(self):
        # Setup
        # deactivate Testsession for this test
        self.session = None
        responses.add(responses.POST,
                      "http://metrics-url/metrics/",
                      json={"foo": "bar"},
                      status=200, content_type='application/json')
        # Execute
        response = self.adminclient.post(
            '/api/metrics/', content_type='application/json')
        # Check
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["scheduled_metrics_initiated"], True)


class TestUserCreation(AuthenticatedAPITestCase):

    def test_create_user_and_token(self):
        # Setup
        user_request = {"email": "test@example.org"}
        # Execute
        request = self.adminclient.post('/api/v1/user/token/', user_request)
        token = request.json().get('token', None)
        # Check
        self.assertIsNotNone(
            token, "Could not receive authentication token on post.")
        self.assertEqual(
            request.status_code, 201,
            "Status code on /api/v1/user/token/ was %s (should be 201)."
            % request.status_code)

    def test_create_user_and_token_fail_nonadmin(self):
        # Setup
        user_request = {"email": "test@example.org"}
        # Execute
        request = self.normalclient.post('/api/v1/user/token/', user_request)
        error = request.json().get('detail', None)
        # Check
        self.assertIsNotNone(
            error, "Could not receive error on post.")
        self.assertEqual(
            error, "You do not have permission to perform this action.",
            "Error message was unexpected: %s."
            % error)

    def test_create_user_and_token_not_created(self):
        # Setup
        user_request = {"email": "test@example.org"}
        # Execute
        request = self.adminclient.post('/api/v1/user/token/', user_request)
        token = request.json().get('token', None)
        # And again, to get the same token
        request2 = self.adminclient.post('/api/v1/user/token/', user_request)
        token2 = request2.json().get('token', None)

        # Check
        self.assertEqual(
            token, token2,
            "Tokens are not equal, should be the same as not recreated.")

    def test_create_user_new_token_nonadmin(self):
        # Setup
        user_request = {"email": "test@example.org"}
        request = self.adminclient.post('/api/v1/user/token/', user_request)
        token = request.json().get('token', None)
        cleanclient = APIClient()
        cleanclient.credentials(HTTP_AUTHORIZATION='Token %s' % token)
        # Execute
        request = cleanclient.post('/api/v1/user/token/', user_request)
        error = request.json().get('detail', None)
        # Check
        # new user should not be admin
        self.assertIsNotNone(
            error, "Could not receive error on post.")
        self.assertEqual(
            error, "You do not have permission to perform this action.",
            "Error message was unexpected: %s."
            % error)


class TestHealthcheckAPI(AuthenticatedAPITestCase):

    def test_healthcheck_read(self):
        # Setup
        # Execute
        response = self.normalclient.get('/api/health/',
                                         content_type='application/json')
        # Check
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["up"], True)
        self.assertEqual(response.data["result"]["database"], "Accessible")


class TestMetrics(AuthenticatedAPITestCase):

    def _check_request(
            self, request, method, params=None, data=None, headers=None):
        self.assertEqual(request.method, method)
        if params is not None:
            url = urlparse.urlparse(request.url)
            qs = urlparse.parse_qsl(url.query)
            self.assertEqual(dict(qs), params)
        if headers is not None:
            for key, value in headers.items():
                self.assertEqual(request.headers[key], value)
        if data is None:
            self.assertEqual(request.body, None)
        else:
            self.assertEqual(json.loads(request.body), data)

    def add_metrics_callback(self):
        responses.add(
            responses.POST, "http://metrics-url/metrics/",
            json={}, status=200, content_type='application/json')

    @responses.activate
    @override_settings(METRICS_AUTH=('metricuser', 'metricpass'))
    def test_direct_fire(self):
        """
        When calling the `fire_metric` task, it should make a POST request
        to the metrics API, with the correct details.
        """
        # Setup
        self.add_metrics_callback()
        # Execute
        result = tasks.fire_metric.apply_async(kwargs={
            "metric_name": 'foo.last',
            "metric_value": 1,
        })
        # Check
        [request] = responses.calls
        self._check_request(
            request.request, 'POST',
            data={"foo.last": 1.0}
        )
        _, auth = request.request.headers['Authorization'].split()
        user, password = b64decode(auth).decode().split(':')
        self.assertEqual(user, 'metricuser')
        self.assertEqual(password, 'metricpass')
        self.assertEqual(result.get(),
                         "Fired metric <foo.last> with value <1.0>")

    @responses.activate
    def test_created_metric(self):
        """
        Ensure that when a new registration is created, all of the relevant
        metrics are fired.
        """
        # Setup
        self.add_metrics_callback()
        # reconnect metric post_save hook
        post_save.connect(fire_created_metric, sender=Registration)

        # Execute
        self.make_registration_adminuser()
        self.make_registration_adminuser()

        # Check
        [request1, request2, request3, request4] = responses.calls
        self._check_request(
            request1.request, 'POST',
            data={"registrations.created.sum": 1.0}
        )
        self._check_request(
            request2.request, 'POST',
            data={"registrations.created.total.last": 1}
        )
        self._check_request(
            request3.request, 'POST',
            data={"registrations.created.sum": 1.0}
        )
        self._check_request(
            request4.request, 'POST',
            data={"registrations.created.total.last": 2}
        )
        # remove post_save hooks to prevent teardown errors
        post_save.disconnect(fire_created_metric, sender=Registration)

    @responses.activate
    def test_source_metric(self):
        # Setup
        self.add_metrics_callback()
        # reconnect metric post_save hook
        post_save.connect(fire_source_metric, sender=Registration)

        # Execute
        self.make_registration_adminuser()

        # Check
        [request] = responses.calls
        self._check_request(
            request.request, 'POST',
            data={"registrations.source.testadminuser.sum": 1.0}
        )
        # remove post_save hooks to prevent teardown errors
        post_save.disconnect(fire_source_metric, sender=Registration)

    @responses.activate
    def test_unique_operator_metric_single(self):
        # Setup
        self.add_metrics_callback()
        # reconnect operator metric post_save hook
        post_save.connect(fire_unique_operator_metric, sender=Registration)

        # Execute
        self.make_registration_adminuser()

        # Check
        [request] = responses.calls
        self._check_request(
            request.request, 'POST',
            data={"registrations.unique_operators.sum": 1.0}
        )

        # Teardown
        # remove post_save hooks to prevent teardown errors
        post_save.disconnect(fire_unique_operator_metric, sender=Registration)

    @responses.activate
    def test_unique_operator_metric_multiple(self):
        # Setup
        # deactivate Testsession for this test
        self.session = None
        # reconnect operator metric post_save hook
        post_save.connect(fire_unique_operator_metric, sender=Registration)
        # prep for a different operator
        new_user_data = {
            "stage": "prebirth",
            "data": REG_DATA['hw_post'],
            "source": self.make_source_adminuser()
        }

        # add metric post response
        responses.add(responses.POST,
                      "http://metrics-url/metrics/",
                      json={"foo": "bar"},
                      status=200, content_type='application/json')

        # Execute
        self.make_registration_adminuser()
        self.make_registration_adminuser()
        self.make_registration_adminuser()
        self.make_registration_adminuser(data=new_user_data)

        # Check
        self.assertEqual(len(responses.calls), 2)
        # remove post_save hooks to prevent teardown errors
        post_save.disconnect(fire_unique_operator_metric, sender=Registration)

    @responses.activate
    def test_message_type_metric(self):
        """
        When creating a registration, two metrics should be fired for the
        message type that the registration is created for, one of type sum, and
        one of type last.
        """
        self.add_metrics_callback()
        post_save.connect(fire_message_type_metric, sender=Registration)

        self.make_registration_adminuser()

        [request_sum, request_last] = responses.calls
        self._check_request(
            request_sum.request, 'POST',
            data={"registrations.msg_type.text.sum": 1.0}
        )
        self._check_request(
            request_last.request, 'POST',
            data={"registrations.msg_type.text.total.last": 1.0}
        )

        post_save.disconnect(fire_message_type_metric, sender=Registration)

    @responses.activate
    def test_receiver_type_metric(self):
        """
        When creating a registration, two metrics should be fired for the
        receiver type that the registration is created for. One of type sum
        with a value of 1, and one of type last with the current total.
        """
        self.add_metrics_callback()
        post_save.connect(fire_receiver_type_metric, sender=Registration)

        self.make_registration_adminuser()

        [request_sum, request_total] = responses.calls
        self._check_request(
            request_sum.request, 'POST',
            data={"registrations.receiver_type.mother_only.sum": 1.0}
        )
        self._check_request(
            request_total.request, 'POST',
            data={"registrations.receiver_type.mother_only.total.last": 1.0}
        )

        post_save.disconnect(fire_receiver_type_metric, sender=Registration)

    @responses.activate
    def test_receiver_type_metric_multiple(self):
        """
        When creating a registration, two metrics should be fired for the
        receiver type that the registration is created for. One of type sum
        with a value of 1, and one of type last with the current total.
        """
        self.add_metrics_callback()
        post_save.connect(fire_receiver_type_metric, sender=Registration)

        cache.clear()
        self.make_registration_adminuser()
        self.make_registration_adminuser()

        [r_sum1, r_total1, r_sum2, r_total2] = responses.calls
        self._check_request(
            r_sum1.request, 'POST',
            data={"registrations.receiver_type.mother_only.sum": 1.0}
        )
        self._check_request(
            r_total1.request, 'POST',
            data={"registrations.receiver_type.mother_only.total.last": 1.0}
        )
        self._check_request(
            r_sum2.request, 'POST',
            data={"registrations.receiver_type.mother_only.sum": 1.0}
        )
        self._check_request(
            r_total2.request, 'POST',
            data={"registrations.receiver_type.mother_only.total.last": 2.0}
        )

        post_save.disconnect(fire_receiver_type_metric, sender=Registration)

    @responses.activate
    def test_message_type_metric_multiple(self):
        """
        When creating a registration, two metrics should be fired for the
        message type that the registration is created for, one of type sum, and
        one of type last. The sum metric should always be one, the last metric
        should increment for each registration of that type.
        """
        self.add_metrics_callback()
        post_save.connect(fire_message_type_metric, sender=Registration)

        cache.clear()
        self.make_registration_adminuser()
        self.make_registration_adminuser()

        [r_sum1, r_last1, r_sum2, r_last2] = responses.calls
        self._check_request(
            r_sum1.request, 'POST',
            data={"registrations.msg_type.text.sum": 1.0}
        )
        self._check_request(
            r_last1.request, 'POST',
            data={"registrations.msg_type.text.total.last": 1.0}
        )
        self._check_request(
            r_sum2.request, 'POST',
            data={"registrations.msg_type.text.sum": 1.0}
        )
        self._check_request(
            r_last2.request, 'POST',
            data={"registrations.msg_type.text.total.last": 2.0}
        )

        post_save.disconnect(fire_message_type_metric, sender=Registration)

    @responses.activate
    def test_language_metric(self):
        """
        When creating a registration, two metrics should be fired for the
        receiver type that the registration is created for. One of type sum
        with a value of 1, and one of type last with the current total.
        """
        self.add_metrics_callback()
        post_save.connect(fire_language_metric, sender=Registration)

        cache.clear()
        self.make_registration_adminuser()
        self.make_registration_adminuser()

        [r_sum1, r_total1, r_sum2, r_total2] = responses.calls
        self._check_request(
            r_sum1.request, 'POST',
            data={"registrations.language.eng_NG.sum": 1.0}
        )
        self._check_request(
            r_total1.request, 'POST',
            data={"registrations.language.eng_NG.total.last": 1.0}
        )
        self._check_request(
            r_sum2.request, 'POST',
            data={"registrations.language.eng_NG.sum": 1.0}
        )
        self._check_request(
            r_total2.request, 'POST',
            data={"registrations.language.eng_NG.total.last": 2.0}
        )

        post_save.disconnect(fire_language_metric, sender=Registration)

    def identity_callback(self, request):
        headers = {'Content-Type': "application/json"}
        resp = {
            "id": "test_id",
            "version": 1,
            "details": {"state": "Abuja", "role": "Midwife"},
            "communicate_through": None,
            "operator": None,
            "created_at": "2016-09-14T17:18:41.629909Z",
            "created_by": 1,
            "updated_at": "2016-09-14T17:18:41.629942Z",
            "updated_by": 1
        }
        return (200, headers, json.dumps(resp))

    def identity_search_callback(self, request):
        headers = {'Content-Type': "application/json"}
        resp = {
            "count": 1,
            "next": None,
            "previous": None,
            "results": [{
                "id": "nurse000-6a07-4377-a4f6-c0485ccba234",
                "version": 1,
                "details": {"state": "Abuja", "role": "Midwife"},
                "communicate_through": None,
                "operator": None,
                "created_at": "2016-09-14T17:18:41.629909Z",
                "created_by": 1,
                "updated_at": "2016-09-14T17:18:41.629942Z",
                "updated_by": 1
            }, ]
        }
        return (200, headers, json.dumps(resp))

    @responses.activate
    def test_state_metric(self):
        """
        When creating a registration, two metrics should be fired for the
        state that the user is registered in. One of type sum with a value of
        1, and one of type last with the current total.
        """
        self.add_metrics_callback()
        post_save.connect(fire_state_metric, sender=Registration)

        operator_id = REG_DATA['hw_pre_mother']['operator_id']

        url = 'http://localhost:8001/api/v1/identities/' + operator_id + "/"
        responses.add_callback(
            responses.GET, url, callback=self.identity_callback,
            content_type="application/json")

        url = 'http://localhost:8001/api/v1/identities/search/?' \
              'details__state=Abuja'
        responses.add_callback(
            responses.GET, url, callback=self.identity_search_callback,
            match_querystring=True, content_type="application/json")

        cache.clear()
        self.make_registration_adminuser()
        self.make_registration_adminuser()

        [_, r_sum1, _, r_total1, _, r_sum2, _, r_total2] = responses.calls
        self._check_request(
            r_sum1.request, 'POST',
            data={"registrations.state.abuja.sum": 1.0}
        )
        self._check_request(
            r_total1.request, 'POST',
            data={"registrations.state.abuja.total.last": 1.0}
        )
        self._check_request(
            r_sum2.request, 'POST',
            data={"registrations.state.abuja.sum": 1.0}
        )
        self._check_request(
            r_total2.request, 'POST',
            data={"registrations.state.abuja.total.last": 2.0}
        )

        post_save.disconnect(fire_state_metric, sender=Registration)

    @responses.activate
    def test_role_metric(self):
        """
        When creating a registration, two metrics should be fired for the
        role that the user is registered as. One of type sum with a value of
        1, and one of type last with the current total.
        """
        self.add_metrics_callback()
        post_save.connect(fire_role_metric, sender=Registration)

        operator_id = REG_DATA['hw_pre_mother']['operator_id']

        url = 'http://localhost:8001/api/v1/identities/' + operator_id + "/"
        responses.add_callback(
            responses.GET, url, callback=self.identity_callback,
            content_type="application/json")

        url = 'http://localhost:8001/api/v1/identities/search/?' \
              'details__role=Midwife'
        responses.add_callback(
            responses.GET, url, callback=self.identity_search_callback,
            match_querystring=True, content_type="application/json")

        cache.clear()
        self.make_registration_adminuser()
        self.make_registration_adminuser()

        [_, r_sum1, _, r_total1, _, r_sum2, _, r_total2] = responses.calls
        self._check_request(
            r_sum1.request, 'POST',
            data={"registrations.role.midwife.sum": 1.0}
        )
        self._check_request(
            r_total1.request, 'POST',
            data={"registrations.role.midwife.total.last": 1.0}
        )
        self._check_request(
            r_sum2.request, 'POST',
            data={"registrations.role.midwife.sum": 1.0}
        )
        self._check_request(
            r_total2.request, 'POST',
            data={"registrations.role.midwife.total.last": 2.0}
        )

        post_save.disconnect(fire_role_metric, sender=Registration)


class TestSubscriptionRequestWebhook(AuthenticatedAPITestCase):

    def test_create_webhook(self):
        # Setup
        user = User.objects.get(username='testadminuser')
        post_data = {
            "target": "http://example.com/registration/",
            "event": "subscriptionrequest.added"
        }
        # Execute
        response = self.adminclient.post('/api/v1/webhook/',
                                         json.dumps(post_data),
                                         content_type='application/json')
        # Check
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        d = Hook.objects.last()
        self.assertEqual(d.target, 'http://example.com/registration/')
        self.assertEqual(d.user, user)

    # This test is not working despite the code working fine
    # If you run these same steps below interactively the webhook will fire
    # @responses.activate
    # def test_mother_only_webhook(self):
    #     # Setup
    #     post_save.connect(receiver=model_saved, sender=SubscriptionRequest,
    #                       dispatch_uid='instance-saved-hook')
    #     Hook.objects.create(user=self.adminuser,
    #                         event='subscriptionrequest.added',
    #                         target='http://example.com/registration/')
    #
    #     expected_webhook = {
    #         "hook": {
    #             "target": "http://example.com/registration/",
    #             "event": "subscriptionrequest.added",
    #             "id": 3
    #         },
    #         "data": {
    #             "messageset": 1,
    #             "updated_at": "2016-02-17T07:59:42.831568+00:00",
    #             "identity": "mother00-9d89-4aa6-99ff-13c225365b5d",
    #             "lang": "eng_NG",
    #             "created_at": "2016-02-17T07:59:42.831533+00:00",
    #             "id": "5282ed58-348f-4a54-b1ff-f702e36ec3cc",
    #             "next_sequence_number": 1,
    #             "schedule": 1
    #         }
    #     }
    #     responses.add(
    #         responses.POST,
    #         "http://example.com/registration/",
    #         json.dumps(expected_webhook),
    #         status=200, content_type='application/json')
    #     registration_data = {
    #         "stage": "prebirth",
    #         "data": REG_DATA["hw_pre_mother"].copy(),
    #         "source": self.make_source_adminuser()
    #     }
    #     registration = Registration.objects.create(**registration_data)
    #     # Execute
    #     result = validate_registration.create_subscriptionrequests(
    #         registration)
    #     # Check
    #     self.assertEqual(result, "1 SubscriptionRequest created")
    #     d_mom = SubscriptionRequest.objects.last()
    #     self.assertEqual(d_mom.identity,
    #                      "mother00-9d89-4aa6-99ff-13c225365b5d")
    #     self.assertEqual(d_mom.messageset, 1)
    #     self.assertEqual(d_mom.next_sequence_number, 1)
    #     self.assertEqual(d_mom.lang, "eng_NG")
    #     self.assertEqual(d_mom.schedule, 1)
    #     self.assertEqual(responses.calls[0].request.url,
    #                      "http://example.com/registration/")


class ManagementTaskTestCase(TestCase):

    def setUp(self):
        def has_listeners():
            return post_save.has_listeners(Registration)
        assert has_listeners(), (
            "Registration model has no post_save listeners. Make sure"
            " helpers cleaned up properly in earlier tests.")
        post_save.disconnect(receiver=registration_post_save,
                             sender=Registration)
        post_save.disconnect(receiver=fire_created_metric,
                             sender=Registration)
        post_save.disconnect(receiver=fire_source_metric,
                             sender=Registration)
        post_save.disconnect(receiver=fire_unique_operator_metric,
                             sender=Registration)
        post_save.disconnect(receiver=fire_message_type_metric,
                             sender=Registration)
        post_save.disconnect(receiver=fire_receiver_type_metric,
                             sender=Registration)
        post_save.disconnect(receiver=fire_language_metric,
                             sender=Registration)
        post_save.disconnect(receiver=fire_state_metric,
                             sender=Registration)
        post_save.disconnect(receiver=fire_role_metric,
                             sender=Registration)
        post_save.disconnect(receiver=model_saved,
                             dispatch_uid='instance-saved-hook')
        assert not has_listeners(), (
            "Registration model still has post_save listeners. Make sure"
            " helpers cleaned up properly in earlier tests.")

        # Mock message set api responses
        query_string = '?short_name=prebirth.mother.text.10_42'
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/messageset/%s' % query_string,
            json={
                "count": 1,
                "next": None,
                "previous": None,
                "results": [{
                    "id": 1,
                    "short_name": 'prebirth.mother.text.10_42',
                    "default_schedule": 1
                }]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )

        # mock household messageset lookup
        query_string = '?short_name=prebirth.household.audio.10_42.fri.9_11'
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/messageset/%s' % query_string,
            json={
                "count": 1,
                "next": None,
                "previous": None,
                "results": [{
                    "id": 3,
                    "short_name": 'prebirth.household.audio.10_42.fri.9_11',
                    "default_schedule": 1
                }]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )

        # mock postbirth mother messageset lookup
        query_string = '?short_name=postbirth.mother.text.0_12'
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/messageset/%s' % query_string,
            json={
                "count": 1,
                "next": None,
                "previous": None,
                "results": [{
                    "id": 1,
                    "short_name": 'postbirth.mother.text.0_12',
                    "default_schedule": 1
                }]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )

        # mock mother schedule lookup
        responses.add(
            responses.GET,
            'http://localhost:8005/api/v1/schedule/1/',
            json={"id": 1, "day_of_week": "1,3,5"},
            status=200, content_type='application/json',
        )

        # mock mother MSISDN lookup
        responses.add(
            responses.GET,
            'http://localhost:8001/api/v1/identities/mother00-9d89-4aa6-99ff-13c225365b5d/addresses/msisdn?default=True',  # noqa
            json={
                "count": 1, "next": None, "previous": None,
                "results": [{"address": "+234123"}]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )

        # mock mother welcome SMS send
        responses.add(
            responses.POST,
            'http://localhost:8006/api/v1/outbound/',
            json={"id": 1},
            status=200, content_type='application/json',
        )

        # mock father MSISDN lookup
        responses.add(
            responses.GET,
            'http://localhost:8001/api/v1/identities/father00-73a2-4d89-b045-d52004c025fe/addresses/msisdn?default=True',  # noqa
            json={
                "count": 1, "next": None, "previous": None,
                "results": [{"address": "+234124"}]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )

    def tearDown(self):
        def has_listeners():
            return post_save.has_listeners(Registration)
        assert not has_listeners(), (
            "Registration model still has post_save listeners. Make sure"
            " helpers removed them properly in earlier tests.")
        post_save.connect(receiver=registration_post_save,
                          sender=Registration)
        post_save.connect(receiver=fire_created_metric,
                          sender=Registration)
        post_save.connect(receiver=fire_source_metric,
                          sender=Registration)
        post_save.connect(receiver=fire_unique_operator_metric,
                          sender=Registration)
        post_save.connect(receiver=fire_language_metric,
                          sender=Registration)
        post_save.connect(receiver=fire_state_metric,
                          sender=Registration)
        post_save.connect(receiver=fire_role_metric,
                          sender=Registration)
        post_save.connect(receiver=model_saved,
                          dispatch_uid='instance-saved-hook')

    def mk_hook(self, event, target='https://www.example.com', user=None):
        user = user or self.user1
        return Hook.objects.create(user=user, event=event, target=target)

    def mk_source(self, user=None):
        data = {
            "name": "test_ussd_source_adminuser",
            "authority": "hw_full",
            "user": user or self.user1
        }
        return Source.objects.create(**data)

    def mk_registration_at_week(self, source, week, today=None,
                                dataset='hw_pre_mother'):
        # prepare registration data
        today = today or datetime.now()
        registration_data = {
            "stage": "prebirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": REG_DATA[dataset].copy(),
            "source": source,
        }
        registration_data["data"].update({
            "preg_week": week,
            "last_period_date": (
                today - timedelta(days=(7 * week))).strftime('%Y%m%d')
        })
        return Registration.objects.create(**registration_data)

    def mk_subscription_request(self, registration):
        validate_registration.create_subscriptionrequests(registration)
        return registration.get_subscription_requests().order_by(
            'created_at').last()

    def do_call_command(self, *args, **kwargs):
        stdout = StringIO()
        stderr = StringIO()

        sbm_url = kwargs.setdefault('sbm_url', 'http://example.com')
        sbm_token = kwargs.setdefault('sbm_token', 'a' * 32)

        args = list(args)
        args.extend(['--sbm-url', sbm_url, '--sbm-token', sbm_token])
        call_command(*args,
                     stdout=stdout, stderr=stderr)
        return stdout, stderr

    def load_subscriptions(self, identity, count=0, results=None):
        responses.add(
            responses.GET,
            'http://example.com/subscriptions/?identity=%s' % (identity,),
            json={
                "count": count,
                "next": None,
                "previous": None,
                "results": results or [],
            },
            status=200, content_type='application/json',
            match_querystring=True
        )


class DummyDeliverer(object):
    """
    A dummy deliverer for webhooks, traps any calls it gets and stores
    them in an internal `.calls` list for introspection.
    """
    def __init__(self):
        self.reset()

    def reset(self):
        self.calls = []

    def __call__(self, *args, **kwargs):
        self.calls.append((args, kwargs))

dummy_deliverer = DummyDeliverer()


class FixRegistrationSubscriptionsTest(ManagementTaskTestCase):

    def setUp(self):
        super(FixRegistrationSubscriptionsTest, self).setUp()
        self.user1 = User.objects.create_user('un1', 'email@example.com', 'pw')
        self.user2 = User.objects.create_user('un2', 'email@example.com', 'pw')

    def tearDown(self):
        super(FixRegistrationSubscriptionsTest, self).tearDown()
        dummy_deliverer.reset()

    def test_crash_on_invalid(self):
        src1 = self.mk_source(self.user1)
        reg1 = self.mk_registration_at_week(src1, week=25)

        self.assertRaisesRegexp(
            CommandError,
            ('This registration is not valid. This command only works with '
             'validated registrations'),
            self.do_call_command, 'fix_registration_subscriptions',
            reg1.id.hex)

    def test_crash_without_sbm_url(self):
        src1 = self.mk_source(self.user1)
        reg1 = self.mk_registration_at_week(src1, week=25)
        reg1.validated = True
        reg1.save()

        self.assertRaisesRegexp(
            CommandError,
            ('Please make sure either the STAGE_BASED_MESSAGING_URL '
             'environment variable or --sbm-url is set.'),
            call_command, 'repopulate_subscriptions', registration=reg1.id.hex)

    def test_crash_without_sbm_token(self):
        src1 = self.mk_source(self.user1)
        reg1 = self.mk_registration_at_week(src1, week=25)
        reg1.validated = True
        reg1.save()

        self.assertRaisesRegexp(
            CommandError,
            ('Please make sure either the STAGE_BASED_MESSAGING_TOKEN '
             'environment variable or --sbm-token is set.'),
            call_command, 'repopulate_subscriptions', registration=reg1.id.hex,
            sbm_url="http://example.com")

    @responses.activate
    def test_fix_registration_without_errors(self):
        src1 = self.mk_source(self.user1)

        reg1 = self.mk_registration_at_week(src1, week=10)
        reg1.validated = True
        reg1.save()
        sub1 = self.mk_subscription_request(reg1)

        # This should be correct
        sub1.next_sequence_number = 1
        sub1.save()

        results = [{
            "id": "test_id",
            "messageset": 1,
            "active": True,
            "next_sequence_number": 1,
        }]
        self.load_subscriptions(reg1.mother_id, count=1, results=results)

        stdout, stderr = self.do_call_command(
            'fix_registration_subscriptions', reg1.id.hex)

        self.assertEqual(
            stdout.getvalue().strip(),
            '%s has correct subscription request %s\nNext sequence numbers '
            'match. Taking no action' % (reg1.id, sub1.id))

        # confirm calling with --fix doesn't make changes
        stdout, stderr = self.do_call_command(
            'fix_registration_subscriptions', reg1.id.hex, '--fix')
        self.assertEqual(1, SubscriptionRequest.objects.all().count())

    @responses.activate
    @mock.patch("registrations.tasks.ValidateRegistration.run")
    def test_fix_registration_without_requests_or_subscriptions(
            self, mock_validation):
        src1 = self.mk_source(self.user1)

        reg1 = self.mk_registration_at_week(src1, week=10)
        reg1.validated = True
        reg1.save()

        self.load_subscriptions(reg1.mother_id, count=0)

        stdout, stderr = self.do_call_command(
            'fix_registration_subscriptions', reg1.id.hex)

        self.assertEqual(
            stdout.getvalue().strip(),
            'Registration %s has no subscriptions or subscription requests'
            % reg1.id)

        # confirm calling with --fix calls the validation task
        stdout, stderr = self.do_call_command(
            'fix_registration_subscriptions', reg1.id.hex, '--fix')
        mock_validation.assert_called_once_with(registration_id=reg1.id)

    @responses.activate
    @override_settings(
        HOOK_DELIVERER='registrations.tests.dummy_deliverer')
    def test_fix_registration_with_correct_request_no_subscriptions(self):
        src1 = self.mk_source(self.user1)

        reg1 = self.mk_registration_at_week(src1, week=10)
        reg1.validated = True
        reg1.save()
        sub1 = self.mk_subscription_request(reg1)

        self.load_subscriptions(reg1.mother_id, count=0)

        stdout, stderr = self.do_call_command(
            'fix_registration_subscriptions', reg1.id.hex)

        self.assertEqual(
            stdout.getvalue().strip(),
            ('%s has correct subscription request %s'
             '\n'
             'No subscription found for subscription request %s'
             ) % (reg1.pk, sub1.pk, sub1.pk,))

        # confirm calling with --fix triggers the subscription hook
        hook1 = self.mk_hook('subscriptionrequest.added')
        # Create an extra hook to make sure we're only firing for 1 not for the
        # whole set
        self.mk_hook('subscriptionrequest.removed')

        stdout, stderr = self.do_call_command(
            'fix_registration_subscriptions', reg1.id.hex, '--fix')

        [request] = SubscriptionRequest.objects.all()
        self.assertEqual(1, request.next_sequence_number)
        [webhook_call] = dummy_deliverer.calls
        args, kwargs = webhook_call
        self.assertEqual(args[0], hook1.target)
        self.assertEqual(args[1]['hook']['id'], hook1.pk)
        self.assertEqual(kwargs['hook'], hook1)
        self.assertEqual(kwargs['instance'], sub1)

    @responses.activate
    def test_fix_registration_with_correct_request_no_active_subscriptions(
            self):
        src1 = self.mk_source(self.user1)

        reg1 = self.mk_registration_at_week(src1, week=10)
        reg1.validated = True
        reg1.save()
        sub1 = self.mk_subscription_request(reg1)

        results = [{
            "id": "test_id",
            "messageset": 1,
            "active": False,
            "next_sequence_number": 1,
        }]
        self.load_subscriptions(reg1.mother_id, count=1, results=results)

        stdout, stderr = self.do_call_command(
            'fix_registration_subscriptions', reg1.id.hex)

        self.assertEqual(
            stdout.getvalue().strip(),
            ('%s has correct subscription request %s'
             '\n'
             'No active subscription found for subscription request %s. '
             'Taking no action'
             ) % (reg1.pk, sub1.pk, sub1.pk,))

    @responses.activate
    def test_fix_registration_with_correct_request_wrong_subscriptions(self):
        src1 = self.mk_source(self.user1)

        reg1 = self.mk_registration_at_week(src1, week=10)
        reg1.validated = True
        reg1.save()
        sub1 = self.mk_subscription_request(reg1)

        # This should be correct
        sub1.next_sequence_number = 1
        sub1.save()

        results = [{
            "id": "test_id",
            "messageset": 1,
            "active": True,
            "next_sequence_number": 50,
        }]
        self.load_subscriptions(reg1.mother_id, count=1, results=results)

        stdout, stderr = self.do_call_command(
            'fix_registration_subscriptions', reg1.id.hex)

        self.assertEqual(
            stdout.getvalue().strip(),
            ('%s has correct subscription request %s'
             '\n'
             'Subscription %s has next_sequence_number=%s, should be %s'
             ) % (reg1.pk, sub1.pk, 'test_id', '50',
                  sub1.next_sequence_number,))

        # confirm calling with --fix updates the subscription
        responses.add(
            responses.PATCH,
            'http://example.com/subscriptions/%s/' % ('test_id',),
            json={
                "id": "test_id",
                "messageset": 1,
                "active": True,
                "next_sequence_number": 1,
            },
            status=200, content_type='application/json',
            match_querystring=True
        )
        stdout, stderr = self.do_call_command(
            'fix_registration_subscriptions', reg1.id.hex, '--fix')
        call = responses.calls[-1]
        self.assertEqual(call.request.url,
                         'http://example.com/subscriptions/test_id/')
        self.assertEqual(call.request.body, '{"next_sequence_number": 1}')

    @responses.activate
    @override_settings(
        HOOK_DELIVERER='registrations.tests.dummy_deliverer')
    def test_fix_registration_with_wrong_request_no_subscriptions(self):
        """
        The subscription request should be fixed and the hook fired
        """
        src1 = self.mk_source(self.user1)

        reg1 = self.mk_registration_at_week(src1, week=25)
        reg1.validated = True
        reg1.save()
        sub1 = self.mk_subscription_request(reg1)

        # This is obviously wrong, should be 25
        sub1.next_sequence_number = 1
        sub1.save()

        self.load_subscriptions(reg1.mother_id, count=0)

        stdout, stderr = self.do_call_command(
            'fix_registration_subscriptions', reg1.id.hex)

        self.assertEqual(
            stdout.getvalue().strip(),
            ('%s has "messageset: 1, next_sequence_number: 1, schedule: 1", '
             'should be "messageset: 1, next_sequence_number: 45, schedule: 1"'
             '\n'
             'No subscription found for subscription request %s'
             ) % (sub1.pk, sub1.pk,))

        # confirm calling with --fix fixes everything
        hook1 = self.mk_hook('subscriptionrequest.added')
        # Create an extra hook to make sure we're only firing for 1 not for the
        # whole set
        self.mk_hook('subscriptionrequest.removed')

        stdout, stderr = self.do_call_command(
            'fix_registration_subscriptions', reg1.id.hex, '--fix')

        [request] = SubscriptionRequest.objects.all()
        self.assertEqual(45, request.next_sequence_number)
        [webhook_call] = dummy_deliverer.calls
        args, kwargs = webhook_call
        self.assertEqual(args[0], hook1.target)
        self.assertEqual(args[1]['hook']['id'], hook1.pk)
        self.assertEqual(kwargs['hook'], hook1)
        self.assertEqual(kwargs['instance'], sub1)

    @responses.activate
    @override_settings(
        HOOK_DELIVERER='registrations.tests.dummy_deliverer')
    def test_fix_registration_household_request(self):
        """
        The father subscription request is wrong and the mother and father
        subscription requests don't have subscriptions. The father subscription
        request should be fixed and the hook fired for both.
        """
        src1 = self.mk_source(self.user1)

        reg1 = self.mk_registration_at_week(src1, week=25,
                                            dataset='hw_pre_father')
        reg1.validated = True
        reg1.save()
        sub2 = self.mk_subscription_request(reg1)
        sub1 = SubscriptionRequest.objects.filter(
            identity=reg1.mother_id).last()

        # This is obviously wrong, should be 45
        sub2.next_sequence_number = 1
        sub2.save()

        self.load_subscriptions(reg1.data['receiver_id'], count=0)
        self.load_subscriptions(reg1.mother_id, count=0)

        stdout, stderr = self.do_call_command(
            'fix_registration_subscriptions', reg1.id.hex)

        self.assertIn(
            ('%s has "messageset: 3, next_sequence_number: 1, schedule: 1", '
             'should be "messageset: 3, next_sequence_number: 45, schedule: 1"'
             '\nNo subscription found for subscription request %s')
            % (sub2.pk, sub2.pk),
            stdout.getvalue().strip())
        self.assertIn(
            ('%s has correct subscription request %s\n'
             'No subscription found for subscription request %s'
             ) % (reg1.id, sub1.pk, sub1.pk),
            stdout.getvalue().strip())

        # confirm calling with --fix fixes everything
        hook1 = self.mk_hook('subscriptionrequest.added')

        stdout, stderr = self.do_call_command(
            'fix_registration_subscriptions', reg1.id.hex, '--fix')

        # confirm father subscription request fixed
        [request] = SubscriptionRequest.objects.filter(
            identity=reg1.data['receiver_id'])
        self.assertEqual(45, request.next_sequence_number)

        # confirm hook fired twice with correct data
        expected_subs = [sub1, sub2]
        [webhook_call1, webhook_call2] = dummy_deliverer.calls
        args, kwargs = webhook_call1
        self.assertEqual(args[0], hook1.target)
        self.assertEqual(args[1]['hook']['id'], hook1.pk)
        self.assertEqual(kwargs['hook'], hook1)
        self.assertIn(kwargs['instance'], expected_subs)
        self.assertEqual(kwargs['instance'].next_sequence_number, 45)
        expected_subs.remove(kwargs['instance'])  # We don't expect it again
        args, kwargs = webhook_call2
        self.assertEqual(args[0], hook1.target)
        self.assertEqual(args[1]['hook']['id'], hook1.pk)
        self.assertEqual(kwargs['hook'], hook1)
        self.assertIn(kwargs['instance'], expected_subs)
        self.assertEqual(kwargs['instance'].next_sequence_number, 45)

    @responses.activate
    def test_fix_registration_only_relevant_subscriptions(self):
        """
        Only the relevant subscriptions should be fixed
        """
        src1 = self.mk_source(self.user1)

        reg1 = self.mk_registration_at_week(src1, week=25)
        reg1.validated = True
        reg1.save()
        sub1 = self.mk_subscription_request(reg1)

        # This is obviously wrong, should be 25
        sub1.next_sequence_number = 1
        sub1.save()

        sub2 = SubscriptionRequest.objects.create(
            identity=reg1.mother_id, messageset=5, next_sequence_number=3,
            lang=reg1.data['language'], schedule=1)

        self.load_subscriptions(reg1.mother_id, count=0)

        stdout, stderr = self.do_call_command(
            'fix_registration_subscriptions', reg1.id.hex)

        self.assertIn(sub2, reg1.get_subscription_requests())
        self.assertEqual(
            stdout.getvalue().strip(),
            ('%s has "messageset: 1, next_sequence_number: 1, schedule: 1", '
             'should be "messageset: 1, next_sequence_number: 45, schedule: 1"'
             '\n'
             'No subscription found for subscription request %s'
             ) % (sub1.pk, sub1.pk,))

    @responses.activate
    @mock.patch("registrations.tasks.ValidateRegistration.run")
    def test_fix_old_pre_birth(self, mock_validation):
        src1 = self.mk_source(self.user1)
        reg1 = self.mk_registration_at_week(src1, week=60)
        reg1.validated = True
        reg1.save()

        self.load_subscriptions(reg1.mother_id, count=0)

        stdout, stderr = self.do_call_command(
            'fix_registration_subscriptions', reg1.id.hex)

        self.assertEqual(
            stdout.getvalue().strip(),
            'Registration %s has no subscriptions or subscription requests'
            % reg1.id)

        # confirm registration didn't change because --fix isn't set
        reg1.refresh_from_db
        self.assertNotIn('baby_dob', reg1.data)

        # confirm calling with --fix calls the validation task
        stdout, stderr = self.do_call_command(
            'fix_registration_subscriptions', reg1.id.hex, '--fix')
        mock_validation.assert_called_once_with(registration_id=reg1.id)
        reg1.refresh_from_db()
        self.assertEqual(reg1.stage, 'postbirth')
        self.assertIn('baby_dob', reg1.data)
        est_dob = datetime.now() - timedelta(
            weeks=(60-settings.PREBIRTH_MAX_WEEKS))
        self.assertEqual(reg1.data['baby_dob'], est_dob.strftime('%Y%m%d'))

    @responses.activate
    def test_fix_postbirth(self):
        post_save.connect(receiver=model_saved,
                          dispatch_uid='instance-saved-hook')
        src1 = self.mk_source(self.user1)
        data = {
            "stage": "postbirth",
            "mother_id": "mother00-9d89-4aa6-99ff-13c225365b5d",
            "data": REG_DATA['hw_post'],
            "source": src1
        }
        reg1 = Registration.objects.create(**data)
        reg1.data['receiver_id'] = reg1.mother_id
        reg1.data['baby_dob'] = "20170101"
        reg1.validated = True
        reg1.save()

        results = [{
            "id": "test_id",
            "messageset": 1,
            "active": True,
            "next_sequence_number": 5,
        }]
        self.load_subscriptions(reg1.mother_id, count=1, results=results)

        stdout, stderr = self.do_call_command(
            'fix_registration_subscriptions', reg1.id.hex, '--today',
            '20170115')

        # Confirm no subscription requests were created
        self.assertEqual(SubscriptionRequest.objects.all().count(), 0)
        self.assertEqual(
            stdout.getvalue().strip(),
            '%s has no subscription requests for postbirth.mother.text.0_12\n'
            'Subscription test_id has next_sequence_number=5, should be 6'
            % reg1.mother_id)

        # confirm calling with --fix fixes things
        responses.add(
            responses.PATCH,
            'http://example.com/subscriptions/%s/' % ('test_id',),
            json={
                "id": "test_id",
                "messageset": 1,
                "active": True,
                "next_sequence_number": 6,
            },
            status=200, content_type='application/json',
            match_querystring=True
        )
        stdout, stderr = self.do_call_command(
            'fix_registration_subscriptions', reg1.id.hex, '--fix', '--today',
            '20170115')
        self.assertEqual(SubscriptionRequest.objects.all().count(), 1)
        call = responses.calls[-1]
        self.assertEqual(call.request.url,
                         'http://example.com/subscriptions/test_id/')
        self.assertEqual(call.request.body, '{"next_sequence_number": 6}')
        post_save.disconnect(receiver=model_saved,
                             dispatch_uid='instance-saved-hook')


class TestRepopulateMetricsTask(TestCase):
    @mock.patch('registrations.tasks.pika')
    @mock.patch('registrations.tasks.RepopulateMetrics.generate_and_send')
    def test_run_repopulate_metrics(self, mock_repopulate, mock_pika):
        """
        The repopulate metrics task should create an amqp connection, and call
        generate_and_send with the appropriate parameters.
        """
        repopulate_metrics.delay(
            'amqp://test', 'prefix', ['metric.foo', 'metric.bar'], '30s:1m')
        args = [args for args, _ in mock_repopulate.call_args_list]

        # Relative instead of absolute times
        start = min(args, key=lambda a: a[3])[3]
        args = [[a, p, m, s-start, e-start] for a, p, m, s, e in args]

        connection = mock_pika.BlockingConnection.return_value
        channel = connection.channel.return_value
        expected = [
            [channel, 'prefix', 'metric.foo',
                timedelta(seconds=0), timedelta(seconds=30)],
            [channel, 'prefix', 'metric.foo',
                timedelta(seconds=30), timedelta(seconds=60)],
            [channel, 'prefix', 'metric.bar',
                timedelta(seconds=0), timedelta(seconds=30)],
            [channel, 'prefix', 'metric.bar',
                timedelta(seconds=30), timedelta(seconds=60)],
        ]

        self.assertEqual(sorted(expected), sorted(args))

        # Assert that the amqp parameters were set from the correc url
        [url], _ = mock_pika.URLParameters.call_args
        self.assertEqual(url, 'amqp://test')
        # Assert that the connection was created with the generated parameters
        [parameters], _ = mock_pika.BlockingConnection.call_args
        self.assertEqual(parameters, mock_pika.URLParameters.return_value)

    @mock.patch('registrations.tasks.MetricGenerator.generate_metric')
    @mock.patch('registrations.tasks.send_metric')
    def test_generate_and_send(
            self, mock_send_metric, mock_metric_generator):
        """
        The generate_and_send function should use the metric generator to
        generate the appropriate metric, then send that metric to Graphite.
        """
        mock_metric_generator.return_value = 17.2
        repopulate_metrics.generate_and_send(
            'amqp://foo', 'prefix', 'foo.bar',
            datetime.utcfromtimestamp(300.0), datetime.utcfromtimestamp(500.0))

        mock_metric_generator.assert_called_once_with(
            'foo.bar', datetime.utcfromtimestamp(300),
            datetime.utcfromtimestamp(500))
        mock_send_metric.assert_called_once_with(
            'amqp://foo', 'prefix', 'foo.bar', 17.2,
            datetime.utcfromtimestamp(400))

    @mock.patch('registrations.tasks.MetricGenerator.generate_metric')
    @mock.patch('registrations.tasks.send_metric')
    def test_generate_and_send_connection_error(
            self, mock_send_metric, mock_metric_generator):
        """
        If the generate_metric function tries to contact an external service
        and fails, then we should just skip sending that metric, and not raise
        an exception.
        """
        mock_metric_generator.side_effect = ConnectTimeout()

        repopulate_metrics.generate_and_send(
            'amqp://foo', 'prefix', 'foo.bar',
            datetime.utcfromtimestamp(300.0), datetime.utcfromtimestamp(500.0))

        mock_metric_generator.assert_called_once_with(
            'foo.bar', datetime.utcfromtimestamp(300),
            datetime.utcfromtimestamp(500))
        mock_send_metric.assert_not_called()

    @mock.patch('registrations.tasks.MetricGenerator.generate_metric')
    @mock.patch('registrations.tasks.send_metric')
    def test_generate_and_send_invalid_json(
            self, mock_send_metric, mock_metric_generator):
        """
        If an external service returns invalid JSON to us, we should just skip
        that specific metric, and not raise an exception.
        """
        mock_metric_generator.side_effect = ValueError()

        repopulate_metrics.generate_and_send(
            'amqp://foo', 'prefix', 'foo.bar',
            datetime.utcfromtimestamp(300.0), datetime.utcfromtimestamp(500.0))

        mock_metric_generator.assert_called_once_with(
            'foo.bar', datetime.utcfromtimestamp(300),
            datetime.utcfromtimestamp(500))
        mock_send_metric.assert_not_called()


def override_get_data(self):
    return [{
        "mothers_phone_number": "07031221927",
        "health_worker_phone_number": "11111",
        "pregnancy_week": "13",
        "gravida": "2",
        "preferred_msg_language": "english",
        "type_of_registration": "prebirth",
        "preferred_msg_type": "voice",
        "message_days": "tuesday_and_thursday",
        "message_time": "2-5pm",
        "message_receiver": "mother_and_father",
        "gatekeeper_phone_number": "07031221928"
    }]


def override_get_data_mother_only(self):
    return [{
        "mothers_phone_number": "07031221927",
        "health_worker_phone_number": "11111",
        "pregnancy_week": "13",
        "gravida": "2",
        "preferred_msg_language": "english",
        "type_of_registration": "prebirth",
        "preferred_msg_type": "voice",
        "message_days": "tuesday_and_thursday",
        "message_time": "2-5pm",
        "message_receiver": "mother_only",
        "gatekeeper_phone_number": None
    }]


def override_get_data_father_only(self):
    return [{
        "mothers_phone_number": "",
        "health_worker_phone_number": "11111",
        "pregnancy_week": "13",
        "gravida": "2",
        "preferred_msg_language": "english",
        "type_of_registration": "prebirth",
        "preferred_msg_type": "voice",
        "message_days": "tuesday_and_thursday",
        "message_time": "2-5pm",
        "message_receiver": "father_only",
        "gatekeeper_phone_number": "07031221927"
    }]


def override_get_data_bad(self):
    return [{
        "mothers_phone_number": "",
        "health_worker_phone_number": "11111",
        "pregnancy_week": "13",
        "gravida": "2",
        "preferred_msg_language": "english",
        "type_of_registration": "prebirth",
        "preferred_msg_type": "voice",
        "message_days": "tuesday_and_thursday",
        "message_time": "2-5pm",
        "message_receiver": "",
        "gatekeeper_phone_number": ""
    }]


class TestThirdPartyRegistrations(AuthenticatedAPITestCase):

    def setUp(self):
        super(TestThirdPartyRegistrations, self).setUp()
        self.orig_get_data = tasks.PullThirdPartyRegistrations.get_data

    def tearDown(self):
        super(TestThirdPartyRegistrations, self).tearDown()
        tasks.PullThirdPartyRegistrations.get_data = self.orig_get_data

    def mock_identity_lookup(self, msisdn, identity_id):
        responses.add(
            responses.GET,
            'http://localhost:8001/api/v1/identities/search/?details__addresses__msisdn=%s' % msisdn,  # noqa
            json={
                "count": 1, "next": None, "previous": None,
                "results": [{
                    "id": identity_id,
                    "details": {}
                }]
            },
            status=200, content_type='application/json',
            match_querystring=True
        )

    def mock_identity_patch(self, identity_id):
        responses.add(
            responses.PATCH,
            'http://localhost:8001/api/v1/identities/%s/' % identity_id,
            json={
                "count": 1, "next": None, "previous": None,
                "results": [{
                    "id": identity_id,
                    "details": {}
                }]
            },

        )

    def mock_identity_post(self, identity_id, communicate_through=None):
        responses.add(
            responses.POST,
            'http://localhost:8001/api/v1/identities/',
            json={
                "id": identity_id,
                "details": {},
                "communicate_through": communicate_through
            },

        )

    def mock_excel_download(self):

        wb = openpyxl.Workbook()
        ws = wb.create_sheet("Forms")

        data = override_get_data(None)[0]

        col = 1
        for key, value in data.items():
            d = ws.cell(row=1, column=col)
            d.value = key

            d = ws.cell(row=2, column=col)
            d.value = value

            col += 1

        responses.add(
            responses.GET,
            settings.THIRDPARTY_REGISTRATIONS_URL,
            status=200,
            body=save_virtual_workbook(wb)
        )

    @responses.activate
    @override_settings(THIRDPARTY_REGISTRATIONS_URL='http://www.3rd.org/d/99/')
    def test_get_data(self):

        self.mock_excel_download()

        task = tasks.PullThirdPartyRegistrations()
        data = task.get_data()

        self.assertEqual(data, override_get_data(None))

    @responses.activate
    def test_start_pull_task(self):
        tasks.PullThirdPartyRegistrations.get_data = override_get_data
        self.make_source_adminuser()

        mother_id = "4038a518-2940-4b15-9c5c-2b7b123b8735"
        father_id = "4038a518-2940-4b15-9c5c-829385793255"
        operator_id = "4038a518-1111-1111-1111-hfud7383gfyt"

        self.mock_identity_lookup("%2B2347031221927", mother_id)
        self.mock_identity_lookup("%2B2347031221928", father_id)
        self.mock_identity_lookup("11111", operator_id)

        self.mock_identity_patch(mother_id)
        self.mock_identity_patch(father_id)

        response = self.adminclient.post('/api/v1/extregistration/',
                                         content_type='application/json')
        # Check
        self.assertEqual(response.status_code,
                         status.HTTP_201_CREATED)

        r = Registration.objects.last()
        self.assertEqual(r.mother_id, "4038a518-2940-4b15-9c5c-2b7b123b8735")
        self.assertEqual(r.stage, "prebirth")
        self.assertEqual(r.data["msg_receiver"], "mother_father")
        self.assertEqual(r.data["operator_id"], operator_id)
        self.assertEqual(r.data["language"], "eng_NG")
        self.assertEqual(r.data["msg_type"], "audio")
        self.assertEqual(r.data["receiver_id"], father_id)
        self.assertEqual(r.data["voice_times"], "2_5")
        self.assertEqual(r.data["voice_days"], "tue_thu")
        self.assertEqual(r.data["last_period_date"], "20150518")

        self.assertEqual(len(responses.calls), 5)
        patch_father = responses.calls[3].request.body
        patch_mother = responses.calls[4].request.body
        self.assertEqual(
            json.loads(patch_father),
            {
                "operator_id": "4038a518-1111-1111-1111-hfud7383gfyt",
                "gravida": "2",
                "preferred_language": "eng_NG",
                "preferred_msg_days": "tue_thu",
                "preferred_msg_type": "audio",
                "household_msgs_only": True,
                "receiver_role": "father",
                "default_addr_type": "msisdn",
                "linked_to": "4038a518-2940-4b15-9c5c-2b7b123b8735",
                "preferred_msg_times": "2_5"
            })
        self.assertEqual(
            json.loads(patch_mother),
            {
                "operator_id": "4038a518-1111-1111-1111-hfud7383gfyt",
                "gravida": "2",
                "preferred_language": "eng_NG",
                "preferred_msg_days": "tue_thu",
                "preferred_msg_type": "audio",
                "receiver_role": "mother",
                "default_addr_type": "msisdn",
                "linked_to": "4038a518-2940-4b15-9c5c-829385793255",
                "preferred_msg_times": "2_5"
            })

    @responses.activate
    def test_start_pull_task_mother_only(self):
        tasks.PullThirdPartyRegistrations.get_data = \
            override_get_data_mother_only
        self.make_source_adminuser()

        mother_id = "4038a518-2940-4b15-9c5c-2b7b123b8735"
        operator_id = "4038a518-1111-1111-1111-hfud7383gfyt"

        self.mock_identity_lookup("%2B2347031221927", mother_id)
        self.mock_identity_lookup("11111", operator_id)

        self.mock_identity_patch(mother_id)

        response = self.adminclient.post('/api/v1/extregistration/',
                                         content_type='application/json')
        # Check
        self.assertEqual(response.status_code,
                         status.HTTP_201_CREATED)

        r = Registration.objects.last()
        self.assertEqual(r.mother_id, "4038a518-2940-4b15-9c5c-2b7b123b8735")
        self.assertEqual(r.stage, "prebirth")
        self.assertEqual(r.data["msg_receiver"], "mother_only")
        self.assertEqual(r.data["operator_id"], operator_id)
        self.assertEqual(r.data["language"], "eng_NG")
        self.assertEqual(r.data["msg_type"], "audio")
        self.assertEqual(r.data["receiver_id"], mother_id)
        self.assertEqual(r.data["voice_times"], "2_5")
        self.assertEqual(r.data["voice_days"], "tue_thu")
        self.assertEqual(r.data["last_period_date"], "20150518")

        self.assertEqual(len(responses.calls), 3)
        patch_mother = responses.calls[2].request.body
        self.assertEqual(
            json.loads(patch_mother),
            {
                "operator_id": "4038a518-1111-1111-1111-hfud7383gfyt",
                "gravida": "2",
                "preferred_language": "eng_NG",
                "preferred_msg_days": "tue_thu",
                "preferred_msg_type": "audio",
                "receiver_role": "mother",
                "default_addr_type": "msisdn",
                "preferred_msg_times": "2_5"
            })

    @responses.activate
    def test_start_pull_task_error(self):
        tasks.PullThirdPartyRegistrations.get_data = override_get_data_bad
        self.make_source_adminuser()

        operator_identity = "4038a518-1111-1111-1111-hfud7383gfyt"

        self.mock_identity_lookup("11111", operator_identity)
        self.mock_identity_patch(operator_identity)

        try:
            self.adminclient.post('/api/v1/extregistration/',
                                  content_type='application/json')
        except:
            pass

        # Check
        self.assertEqual(Registration.objects.count(), 0)

        e = ThirdPartyRegistrationError.objects.last()
        self.assertTrue(e.data['error'].find("mother_id") != -1)
        self.assertTrue(
            e.data['error'].find("This field may not be null.") != -1)

    @responses.activate
    def test_start_pull_task_connection_error(self):
        tasks.PullThirdPartyRegistrations.get_data = override_get_data
        self.make_source_adminuser()

        operator_identity = "4038a518-1111-1111-1111-hfud7383gfyt"

        self.mock_identity_lookup("11111", operator_identity)
        self.mock_identity_patch(operator_identity)

        try:
            self.adminclient.post('/api/v1/extregistration/',
                                  content_type='application/json')
        except:
            pass

        # Check
        self.assertEqual(Registration.objects.count(), 0)

        e = ThirdPartyRegistrationError.objects.last()

        self.assertTrue(e.data['error'].find("Connection refused") != -1)


class TestAddRegistrationsAPI(TestThirdPartyRegistrations):

    @responses.activate
    def test_add_registration(self):
        """
        It should successfully create a registration if the data is valid
        """

        self.make_source_adminuser()

        mother_id = "4038a518-2940-4b15-9c5c-2b7b123b8735"
        operator_id = "4038a518-1111-1111-1111-hfud7383gfyt"

        self.mock_identity_lookup("%2B2347031221927", mother_id)
        self.mock_identity_lookup("11111", operator_id)

        self.mock_identity_patch(mother_id)

        data = override_get_data_mother_only(None)[0]

        response = self.adminclient.post('/api/v1/addregistration/',
                                         json.dumps(data),
                                         content_type='application/json')

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data['registration_added'], True)

        self.assertEqual(Registration.objects.count(), 1)

        reg = Registration.objects.last()
        self.assertEqual(
            reg.created_by, User.objects.get(username='testadminuser'))
        self.assertEqual(
            reg.updated_by, User.objects.get(username='testadminuser'))

    @responses.activate
    def test_add_registration_father_only(self):
        """
        It should successfully create a father_only registration if the data
        is valid
        """

        self.make_source_adminuser()

        father_id = "4038a518-2940-4b15-9c5c-2b7b123b8735"
        mother_id = "4038a518-2940-4b15-9c5c-skjdfnsjdnff"
        operator_id = "4038a518-1111-1111-1111-hfud7383gfyt"

        self.mock_identity_lookup("%2B2347031221927", father_id)
        self.mock_identity_lookup("11111", operator_id)
        self.mock_identity_post(mother_id, father_id)

        self.mock_identity_patch(mother_id)
        self.mock_identity_patch(father_id)

        data = override_get_data_father_only(None)[0]

        response = self.adminclient.post('/api/v1/addregistration/',
                                         json.dumps(data),
                                         content_type='application/json')

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data['registration_added'], True)

        self.assertEqual(Registration.objects.count(), 1)

        reg = Registration.objects.last()
        self.assertEqual(reg.mother_id, mother_id)
        self.assertEqual(reg.data['receiver_id'], father_id)
        self.assertEqual(
            reg.created_by, User.objects.get(username='testadminuser'))
        self.assertEqual(
            reg.updated_by, User.objects.get(username='testadminuser'))

    @responses.activate
    def test_add_registration_no_source(self):
        """
        It should not create a registration if the user doesn't have a source
        even if the data is valid
        """

        mother_id = "4038a518-2940-4b15-9c5c-2b7b123b8735"
        operator_id = "4038a518-1111-1111-1111-hfud7383gfyt"

        self.mock_identity_lookup("%2B2347031221927", mother_id)
        self.mock_identity_lookup("11111", operator_id)

        self.mock_identity_patch(mother_id)

        data = override_get_data_mother_only(None)[0]

        response = self.adminclient.post('/api/v1/addregistration/',
                                         json.dumps(data),
                                         content_type='application/json')

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data['registration_added'], False)
        self.assertEqual(response.data['error'],
                         "Source matching query does not exist.")

        self.assertEqual(Registration.objects.count(), 0)

    @responses.activate
    def test_add_registration_missing_field(self):
        """
        It should not create a registration if there are missing fields, it
        should repond with correct code and message
        """

        self.make_source_adminuser()

        mother_id = "4038a518-2940-4b15-9c5c-2b7b123b8735"
        operator_id = "4038a518-1111-1111-1111-hfud7383gfyt"

        self.mock_identity_lookup("%2B2347031221927", mother_id)
        self.mock_identity_lookup("11111", operator_id)

        self.mock_identity_patch(mother_id)

        data = override_get_data_mother_only(None)[0]
        del data['gravida']

        response = self.adminclient.post('/api/v1/addregistration/',
                                         json.dumps(data),
                                         content_type='application/json')

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data['registration_added'], False)
        self.assertEqual(response.data['error'], "Missing field: 'gravida'")

        self.assertEqual(Registration.objects.count(), 0)

    @responses.activate
    def test_add_registration_invalid_data(self):
        """
        It should not create a registration if there is invalid data, it
        should repond with correct code and message
        """

        self.make_source_adminuser()

        mother_id = "4038a518-2940-4b15-9c5c-2b7b123b8735"
        operator_id = "4038a518-1111-1111-1111-hfud7383gfyt"

        self.mock_identity_lookup("%2B2347031221927", mother_id)
        self.mock_identity_lookup("11111", operator_id)

        self.mock_identity_patch(mother_id)

        data = override_get_data_bad(None)[0]

        response = self.adminclient.post('/api/v1/addregistration/',
                                         json.dumps(data),
                                         content_type='application/json')

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data['registration_added'], False)
        self.assertTrue(
            response.data['error'].find("This field may not be null") != -1)

        self.assertEqual(Registration.objects.count(), 0)
